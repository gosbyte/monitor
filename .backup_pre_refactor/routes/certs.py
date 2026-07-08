# -*- coding: utf-8 -*-
"""到期项 CRUD 路由 - 增删改查/导入导出/备份恢复"""
from __future__ import annotations

import csv
import io
import json
import os
import re
import time
import secrets
import logging
from datetime import datetime, timedelta
from typing import Any, Union

from flask import Flask, request, jsonify, render_template, redirect, url_for, Response, session

from exceptions import ValidationError, DataError, ServiceError, ImportError, ExportError

from data import (
    load_certs, save_certs, load_config, save_config, write_log,
    calc_days_left, get_cert_status, calc_stats, encrypt_field, decrypt_field,
    load_users, DATA_DIR, BASE_DIR, DATA_FILE, CONFIG_FILE, LOGS_FILE,
    USERS_FILE, USE_SQLITE,
)
from auth import login_required, csrf_required, admin_required


# Flask route handlers can return str, tuple[str, int], Response, or dict
_FlaskResponse = Union[str, tuple[str, int], Response, dict[str, Any], Any]

logger = logging.getLogger(__name__)


def _check_api_csrf() -> bool:
    """API CSRF 检查（供蓝图内部使用）"""
    if request.method == "GET":
        return True
    token = request.headers.get("X-CSRF-Token")
    if not token and request.is_json:
        token = request.json.get("_csrf_token")  # type: ignore[union-attr]
    if not token or token != session.get("_csrf_token"):
        return False
    if request.method in ("POST", "PUT", "DELETE", "PATCH"):
        session["_csrf_token"] = secrets.token_hex(32)
    return True


def register_cert_routes(app: Flask) -> None:
    """注册到期项相关路由"""

    # ── 仪表盘 ────────────────────────────────────────────────
    @app.route("/")
    @login_required
    def index() -> _FlaskResponse:
        username = session.get("username", "")
        users = load_users()
        current_user = next((u for u in users if u["username"] == username), None)
        if current_user and current_user.get("force_change_password", 0):
            return redirect(url_for("change_password"))

        certs = load_certs()
        cfg = load_config()
        for c in certs:
            # [FIX] 统一日期格式：确保 expire_date 格式正确
            ed = c.get("expire_date", "")
            if ed and "T" in ed:
                c["expire_date"] = ed.replace("T", " ")
            c["days_left"] = calc_days_left(c["expire_date"])
            c["status"] = get_cert_status(c, c["days_left"])
        certs.sort(key=lambda x: x["days_left"])
        stats = calc_stats(certs)
        users = load_users()
        current_username = session.get("username", "")
        current_user = next((u for u in users if u["username"] == current_username), None)
        is_admin = current_user.get("role") == "admin" if current_user else False
        if not is_admin:
            certs = [c for c in certs if c.get("created_by") == current_username]

        cert_types = sorted(set(c["cert_type"] for c in certs if c.get("cert_type")))

        from collections import defaultdict
        monthly_count = defaultdict(int)
        today = datetime.now()
        for c in certs:
            days_left = calc_days_left(c["expire_date"])
            if days_left >= 0:
                expire_dt = today + timedelta(days=int(days_left))
                month_key = expire_dt.strftime("%Y-%m")
                monthly_count[month_key] += 1
        monthly_expiry = []
        for i in range(6):
            m_month = today.month + i
            m_year = today.year + (m_month - 1) // 12
            m_month = (m_month - 1) % 12 + 1
            m_key = f"{m_year}-{m_month:02d}"
            monthly_expiry.append({"month": m_key, "count": monthly_count.get(m_key, 0)})
        max_monthly = max([m["count"] for m in monthly_expiry]) if monthly_expiry else 0

        type_count = defaultdict(int)
        for c in certs:
            t = c.get("cert_type", "其他")
            type_count[t] += 1
        total_certs = len(certs) if len(certs) > 0 else 1
        type_distribution = [{"type": t, "count": cnt, "percent": round(cnt*100/total_certs, 1)} for t, cnt in sorted(type_count.items(), key=lambda x: -x[1])[:8]]

        status_distribution = [
            {"label": "正常", "count": stats["normal"], "color": "#22c55e"},
            {"label": "即将到期", "count": stats["expiring"], "color": "#f97316"},
            {"label": "已过期", "count": stats["expired"], "color": "#ef4444"},
            {"label": "已禁用", "count": stats.get("disabled", 0), "color": "#6b7280"}
        ]

        # [FIX] P1-9: badge_count 在此处计算
        badge_count = sum(1 for c in certs if c.get("remind_enabled", True) and not c.get("handled", False) and 0 <= c.get("days_left", 999) <= 7)

        chart_data: dict[str, Any] = {
            "monthly_expiry": monthly_expiry,
            "max_monthly": max_monthly,
            "type_distribution": type_distribution,
            "status_distribution": status_distribution
        }
        return render_template("index.html", certs=certs, cfg=cfg, stats=stats, users=users, is_admin=is_admin,
                               chart_data=chart_data, cert_types=cert_types, current_username=current_username,
                               badge_count=badge_count, active_page='index', page_title='到期提醒管理系统',
                               csrf_token=session.get("_csrf_token", ""))

    @app.route("/add", methods=["POST"])
    @login_required
    @csrf_required
    def add_cert() -> _FlaskResponse:
        is_ajax = request.headers.get("X-Requested-With") == "XMLHttpRequest" or request.form.get("_ajax") == "1"
        certs = load_certs()

        # [FIX] P0-3: 添加到期项
        from db import db_save_cert, get_db
        cert_data: dict[str, Any] = {
            "customer": request.form.get("customer", "").strip(),
            "cert_type": (request.form.get("cert_type", "").strip() or request.form.get("cert_type_custom", "").strip()),
            "domain": request.form.get("domain", "").strip(),
            "expire_date": (request.form.get("expire_date", "").strip() + " " + request.form.get("expire_time", "").strip()).strip(),
            "note": request.form.get("note", "").strip(),
            "remind_enabled": request.form.get("remind_enabled", "on") == "on",
            "handled": False,
            "responsible_users": request.form.getlist("responsible_users"),
            "created_by": session.get("username", ""),
            "created_at": datetime.now().strftime("%Y-%m-%d %H:%M")
        }
        db_save_cert(cert_data)
        # 获取新 ID
        with get_db() as conn:
            row = conn.execute("SELECT MAX(id) as max_id FROM certs").fetchone()
            new_id = row[0] if row[0] else 1
        customer = cert_data["customer"]

        # [FIX] P0-6: write_log 参数正确传递
        write_log(session.get("username", "?"), "添加记录", customer, "到期项", request.remote_addr or '')

        if is_ajax:
            # Calculate days_left for the new cert
            expire_dt = datetime.strptime(cert_data["expire_date"], "%Y-%m-%d %H:%M")
            now_dt = datetime.now()
            days_left = (expire_dt - now_dt).total_seconds() / 86400
            status = "expired" if days_left < 0 else ("expiring" if days_left <= 7 else ("warning" if days_left <= 30 else "normal"))
            if cert_data.get("remind_enabled") == False:
                status = "disabled"
            new_cert = {
                "id": new_id,
                "customer": cert_data["customer"],
                "cert_type": cert_data["cert_type"],
                "expire_date": cert_data["expire_date"],
                "note": cert_data["note"] or "",
                "days_left": round(days_left, 1),
                "status": status,
                "handled": False,
                "remind_enabled": cert_data.get("remind_enabled", True),
                "created_by": cert_data.get("created_by", ""),
            }
            return jsonify(ok=True, id=new_id, message="添加成功", new_cert=new_cert, csrf_token=session.get("_csrf_token", ""))
        return redirect(url_for("index") + "?success=添加成功")

    @app.route("/edit/<int:cert_id>", methods=["GET", "POST"])
    @login_required
    @csrf_required
    def edit_cert(cert_id: int) -> _FlaskResponse:
        users = load_users()
        current_username = session.get("username", "")
        current_user = next((u for u in users if u["username"] == current_username), None)
        is_admin = current_user.get("role") == "admin" if current_user else False
        certs = load_certs()
        cert = next((c for c in certs if c["id"] == cert_id), None)
        if not cert:
            raise ValidationError("记录不存在")
        if cert.get("created_by") and cert["created_by"] != current_username and not is_admin:
            raise DataError("无权操作此记录")
        if request.method == "POST":
            is_ajax = request.is_json or request.headers.get("Content-Type", "").startswith("application/json")
            if is_ajax:
                data = request.get_json()
                cert["customer"] = data.get("customer", "").strip()
                cert["cert_type"] = data.get("cert_type", "").strip()
                cert["expire_date"] = data.get("expire_date", "").strip()
                cert["note"] = data.get("note", "").strip()
                cert["remind_enabled"] = bool(data.get("remind_enabled", True))
                cert["handled"] = bool(data.get("handled", False))
                cert["responsible_users"] = data.get("responsible_users", [])
            else:
                ct = request.form.get("cert_type", "").strip()
                if not ct:
                    ct = request.form.get("cert_type_custom", "").strip()
                cert["cert_type"] = ct
                cert["customer"] = request.form.get("customer", "").strip()
                cert["domain"] = request.form.get("domain", "").strip()
                cert["expire_date"] = request.form.get("expire_date", "").strip()
                cert["note"] = request.form.get("note", "").strip()
                cert["remind_enabled"] = request.form.get("remind_enabled", "on") == "on"
                cert["handled"] = request.form.get("handled") == "on"
                cert["responsible_users"] = request.form.getlist("responsible_users")
            save_certs(certs)
            write_log(session.get("username", "?"), f"编辑记录 #{cert_id}", cert.get("customer", ""), "到期项", request.remote_addr or '')
            if is_ajax:
                return jsonify({"ok": True, "success": True, "csrf_token": session.get("_csrf_token", "")})
            return redirect(url_for("index") + "?success=保存成功")
        return render_template("edit.html", cert=cert, users=users, is_admin=is_admin)

    @app.route("/delete/<int:cert_id>", methods=["POST"])
    @login_required
    @csrf_required
    @admin_required
    def delete_cert(cert_id: int) -> _FlaskResponse:
        is_ajax = request.headers.get("X-Requested-With") == "XMLHttpRequest"
        certs = load_certs()
        cert_to_delete = next((c for c in certs if c["id"] == cert_id), None)
        cert_name = cert_to_delete["customer"] if cert_to_delete else str(cert_id)
        certs = [c for c in certs if c["id"] != cert_id]
        save_certs(certs)
        write_log(session.get("username", "?"), f"删除记录 #{cert_id}", cert_name, "到期项", request.remote_addr or '')
        if is_ajax:
            return jsonify(ok=True, message="删除成功", csrf_token=session.get("_csrf_token", ""))
        return redirect(url_for("index") + "?success=删除成功")

    @app.route("/api/cert/<int:cert_id>", methods=["DELETE"])
    @login_required
    def api_delete_cert(cert_id: int) -> Any:
        if not _check_api_csrf():
            return jsonify({"ok": False, "message": "CSRF验证失败"}), 403
        users = load_users()
        current_username = session.get("username", "")
        current_user = next((u for u in users if u["username"] == current_username), None)
        is_admin = current_user.get("role") == "admin" if current_user else False
        certs = load_certs()
        cert = next((c for c in certs if c["id"] == cert_id), None)
        if not cert:
            return jsonify({"ok": False, "message": "记录不存在"}), 404
        if not is_admin and cert.get("created_by") and cert["created_by"] != current_username:
            return jsonify({"ok": False, "message": "无权删除此记录"}), 403
        cert_name = cert.get("customer", str(cert_id))
        certs = [c for c in certs if c["id"] != cert_id]
        save_certs(certs)
        write_log(current_username, f"删除记录 #{cert_id}", cert_name, "到期项", request.remote_addr or '')
        return jsonify({"ok": True, "message": "删除成功", "csrf_token": session.get("_csrf_token", "")})

    @app.route("/api/cert_status/<int:cert_id>")
    @login_required
    def get_cert_status_api(cert_id: int) -> Any:
        certs = load_certs()
        for c in certs:
            if c["id"] == cert_id:
                days_left = calc_days_left(c["expire_date"])
                status = get_cert_status(c, days_left)
                enabled = c.get("remind_enabled", True)
                handled = c.get("handled", False)
                # 统一日期格式用于显示
                expire_str = c.get("expire_date", "").replace("T", " ").strip()
                if status == "disabled":
                    badge = f'<span class="inline-flex items-center gap-1 px-2 py-1 rounded-full text-xs font-medium bg-gray-100 text-gray-600" title="到期日期：{expire_str}"><i data-lucide="bell-off" class="w-3 h-3"></i> 已禁用</span>'
                elif status == "expired":
                    badge = f'<span class="inline-flex items-center gap-1 px-2 py-1 rounded-full text-xs font-medium bg-red-100 text-red-800" title="到期日期：{expire_str}"><i data-lucide="x" class="w-3 h-3"></i> 已过期 {abs(days_left):.0f}天</span>'
                elif status == "expiring":
                    badge = f'<span class="inline-flex items-center gap-1 px-2 py-1 rounded-full text-xs font-medium bg-orange-100 text-orange-800" title="到期日期：{expire_str}"><i data-lucide="alert-triangle" class="w-3 h-3"></i> {days_left:.0f}天后</span>'
                elif status == "normal":
                    badge = f'<span class="inline-flex items-center gap-1 px-2 py-1 rounded-full text-xs font-medium bg-blue-50 text-blue-700" title="到期日期：{expire_str}"><i data-lucide="check-circle" class="w-3 h-3"></i> {days_left:.0f}天</span>'
                else:
                    badge = f'<span>{expire_str}</span>'
                return jsonify({
                    "ok": True, "days_left": days_left, "status": status, "badge_html": badge,
                    "remind_enabled": enabled, "handled": handled,
                    "responsible_users": c.get("responsible_users", []),
                    "customer": c.get("customer", ""), "cert_type": c.get("cert_type", ""),
                    "expire_date": c.get("expire_date", ""), "note": c.get("note", "")
                })
        return jsonify({"ok": False}), 404

    # ── 批量导入 / 导出 ──────────────────────────────────────
    @app.route("/import", methods=["POST"])
    @login_required
    @admin_required
    @csrf_required
    def import_certs() -> Any:
        certs = load_certs()
        try:
            data = request.get_json()
            if not isinstance(data, list):
                return jsonify({"ok": False, "message": "请传入 JSON 数组格式"}), 400
            imported = 0
            errors: list[str] = []
            new_id = max([c["id"] for c in certs], default=0)
            for i, item in enumerate(data):
                try:
                    customer = str(item.get("customer", "").strip())
                    expire_date = str(item.get("expire_date", "").strip())
                    if not customer or not expire_date:
                        errors.append(f"第 {i+1} 条: 缺少必填字段(客户名/到期日期)")
                        continue
                    new_id += 1
                    certs.append({
                        "id": new_id, "customer": customer,
                        "cert_type": str(item.get("cert_type", "").strip()),
                        "domain": str(item.get("domain", "").strip()),
                        "expire_date": expire_date,
                        "note": str(item.get("note", "").strip()),
                        "remind_enabled": bool(item.get("remind_enabled", True)),
                        "handled": bool(item.get("handled", False)),
                        "responsible_users": item.get("responsible_users", []),
                        "created_by": session.get("username", ""),
                        "created_at": datetime.now().strftime("%Y-%m-%d %H:%M")
                    })
                    imported += 1
                except Exception as e:
                    errors.append(f"第 {i+1} 条: {str(e)}")
            save_certs(certs)
            write_log(session.get("username", "?"), "批量导入", f"共 {imported} 条记录", "到期项", request.remote_addr or '')
            return jsonify({"ok": True, "imported": imported, "errors": errors})
        except Exception as e:
            logger.exception("批量导入失败")
            raise ServiceError(f"导入失败: {str(e)}")

    @app.route("/export")
    @app.route("/export/json")
    @login_required
    @admin_required
    def export_certs() -> Response:
        certs = load_certs()
        for c in certs:
            c["days_left"] = calc_days_left(c["expire_date"])
            c["status"] = get_cert_status(c, c["days_left"])
        return Response(
            json.dumps(certs, ensure_ascii=False, indent=2),
            mimetype="application/json",
            headers={"Content-Disposition": "attachment; filename=cert_data_export.json"}
        )

    @app.route("/export/excel")
    @login_required
    @admin_required
    def export_excel() -> Response:
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "导出数据"
        ws.append(["客户名称", "提醒类型", "域名", "到期日期", "是否提醒", "处理状态", "备注"])
        certs = load_certs()
        for c in certs:
            ws.append([
                c.get("customer", ""), c.get("cert_type", ""), c.get("domain", ""),
                c.get("expire_date", ""),
                "是" if c.get("remind_enabled", True) else "否",
                "已处理" if c.get("handled", False) else "未处理",
                c.get("note", "")
            ])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return Response(
            buf.read(),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=cert_data_export.xlsx"}
        )

    @app.route("/import/template")
    @login_required
    @admin_required
    def download_template() -> Response:
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "导入模板"
        ws.append(["客户名称", "提醒类型", "域名", "到期日期", "备注"])
        ws.append(["示例客户", "SSL到期项", "example.com", "2026-12-31", "备注信息"])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return Response(
            buf.read(),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=import_template.xlsx"}
        )

    @app.route("/api/preview_import", methods=["POST"])
    @login_required
    @admin_required
    def api_preview_import() -> Any:
        if not _check_api_csrf():
            return jsonify({"ok": False, "message": "CSRF验证失败", "csrf_token": session.get("_csrf_token")})
        if "file" not in request.files:
            return jsonify({"ok": False, "message": "未上传文件", "csrf_token": session.get("_csrf_token")})
        file = request.files["file"]
        if not file.filename.endswith(".xlsx"):
            return jsonify({"ok": False, "message": "只支持 .xlsx 文件", "csrf_token": session.get("_csrf_token")})
        try:
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(file.read()))
            ws = wb.active
            rows = list(ws.iter_rows(min_row=2, values_only=True))
            data: list[dict[str, str]] = []
            for r in rows:
                if not r or not r[0]:
                    continue
                data.append({
                    "customer": str(r[0] or "").strip(),
                    "cert_type": str(r[1] or "").strip(),
                    "domain": str(r[2] or "").strip(),
                    "expiry_date": str(r[3] or "").strip(),
                    "note": str(r[4] or "").strip()
                })
            return jsonify({"ok": True, "data": data, "total": len(data), "csrf_token": session.get("_csrf_token")})
        except Exception as e:
            return jsonify({"ok": False, "message": str(e), "csrf_token": session.get("_csrf_token")})

    @app.route("/api/import_excel", methods=["POST"])
    @login_required
    @admin_required
    def api_import_excel() -> Any:
        if not _check_api_csrf():
            return jsonify({"ok": False, "message": "CSRF验证失败", "csrf_token": session.get("_csrf_token")})
        payload = request.get_json() or {}
        rows = payload.get("data", [])
        if not rows:
            return jsonify({"ok": False, "message": "导入数据为空", "csrf_token": session.get("_csrf_token")})
        certs = load_certs()
        imported = 0
        errors: list[str] = []
        for i, r in enumerate(rows):
            customer = r.get("customer", "").strip()
            expiry = r.get("expiry_date", "").strip()
            if not customer or not expiry:
                errors.append(f"第{i+1}行: 缺少必填字段")
                continue
            try:
                calc_days_left(expiry)
            except Exception:
                errors.append(f"第{i+1}行: 日期格式无效 '{expiry}'")
                continue
            if any(c.get("customer") == customer and c.get("expire_date") == expiry for c in certs):
                errors.append(f"第{i+1}行: 重复记录")
                continue
            new_id = max([c["id"] for c in certs], default=0) + 1
            certs.append({
                "id": new_id, "customer": customer,
                "cert_type": r.get("cert_type", ""), "domain": r.get("domain", ""),
                "expire_date": expiry, "note": r.get("note", ""),
                "remind_enabled": True, "handled": False,
                "created_by": session.get("username", ""),
                "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            })
            imported += 1
        save_certs(certs)
        write_log(session.get("username", "?"), "Excel导入", f"成功导入 {imported} 条，失败 {len(errors)} 条", "到期项", request.remote_addr or '')
        return jsonify({"ok": True, "message": f"成功导入 {imported} 条记录", "imported": imported, "errors": errors, "csrf_token": session.get("_csrf_token")})

    # ── 备份 / 恢复 ──────────────────────────────────────────
    @app.route("/backup")
    @login_required
    @admin_required
    def backup_data() -> Response:
        from db import db_load_certs, db_load_logs, db_load_push_history, db_load_config, get_db
        certs = db_load_certs()
        logs = db_load_logs()
        push_history = db_load_push_history()
        cfg = db_load_config()
        users: list[dict[str, Any]] = []
        with get_db() as conn:
            rows = conn.execute("SELECT * FROM users").fetchall()
            users = [dict(r) for r in rows]
        backup: dict[str, Any] = {
            "backup_time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "version": "2.0", "mode": "sqlite",
            "cert_data": certs, "config": cfg, "users": users,
            "logs": logs, "push_history": push_history,
        }
        filename = "backup_" + datetime.now().strftime("%Y%m%d_%H%M%S") + ".json"
        return Response(
            json.dumps(backup, ensure_ascii=False, indent=2),
            mimetype="application/json",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    @app.route("/restore", methods=["GET", "POST"])
    @login_required
    @admin_required
    def restore_data() -> Any:
        if request.method == "GET":
            return render_template("restore.html")
        if "backup_file" not in request.files:
            return jsonify({"ok": False, "message": "未找到上传文件"})
        file = request.files["backup_file"]
        if file.filename == "":
            return jsonify({"ok": False, "message": "请选择备份文件"})
        try:
            data = json.load(io.TextIOWrapper(file, encoding="utf-8-sig"))
        except Exception as e:
            return jsonify({"ok": False, "message": f"文件格式错误：{e}"})
        try:
            from db import (get_db, db_save_cert, db_load_users, db_save_user,
                            db_load_config, db_save_config, db_write_log,
                            db_save_push_history)
            if "cert_data" in data and data["cert_data"]:
                with get_db() as conn:
                    for cert in data["cert_data"]:
                        cert_id = cert.get("id", 0)
                        existing = conn.execute("SELECT id FROM certs WHERE id=?", (cert_id,)).fetchone()
                        if existing:
                            conn.execute("""UPDATE certs SET customer=?, cert_type=?, domain=?, expire_date=?,
                                           note=?, remind_enabled=?, handled=?, responsible_users=?, updated_at=?
                                           WHERE id=?""",
                                (cert.get("customer", ""), cert.get("cert_type", ""), cert.get("domain", ""),
                                 cert.get("expire_date", ""), cert.get("note", ""),
                                 int(cert.get("remind_enabled", True)), int(cert.get("handled", False)),
                                 json.dumps(cert.get("responsible_users", []), ensure_ascii=False),
                                 datetime.now().strftime("%Y-%m-%d %H:%M"), cert_id))
                        else:
                            conn.execute("""INSERT OR REPLACE INTO certs (id, customer, cert_type, domain, expire_date, note,
                                          remind_enabled, handled, responsible_users, created_by, created_at, updated_at)
                                          VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                                (cert_id, cert.get("customer", ""), cert.get("cert_type", ""), cert.get("domain", ""),
                                 cert.get("expire_date", ""), cert.get("note", ""),
                                 int(cert.get("remind_enabled", True)), int(cert.get("handled", False)),
                                 json.dumps(cert.get("responsible_users", []), ensure_ascii=False),
                                 cert.get("created_by", ""), cert.get("created_at", ""), datetime.now().strftime("%Y-%m-%d %H:%M")))
            if "config" in data and data["config"]:
                db_save_config(data["config"])
            if "users" in data and data["users"]:
                existing_users = {u["username"] for u in db_load_users()}
                for user in data["users"]:
                    uname = user.get("username", "")
                    if uname in existing_users:
                        db_save_user(user)
                    else:
                        with get_db() as conn:
                            conn.execute("""INSERT INTO users (username, name, password, dingtalk_id,
                                           role, email, failed_attempts, consecutive_locks, lock_until,
                                           force_change_password)
                                           VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                                (uname, user.get("name", uname), user.get("password", ""),
                                 user.get("dingtalk_id", ""), user.get("role", "user"),
                                 user.get("email", ""), user.get("failed_attempts", 0),
                                 user.get("consecutive_locks", 0), user.get("lock_until"),
                                 int(user.get("force_change_password", 1))))
            if "logs" in data and data["logs"]:
                for log in data["logs"][-1000:]:
                    db_write_log(log.get("username", ""), log.get("action", ""),
                               log.get("detail", ""), log.get("target", ""), log.get("ip", ""))
            if "push_history" in data and data["push_history"]:
                for ph in data["push_history"]:
                    db_save_push_history(ph.get("cert_customer", ""), ph.get("cert_domain", ""),
                                         ph.get("channels", []), ph.get("status", ""), ph.get("message", ""))
        except Exception as e:
            logger.exception("数据恢复失败")
            raise ServiceError(f"恢复失败：{e}")
        write_log(session["username"], "恢复数据", "系统", f"从备份恢复（{file.filename}）", request.remote_addr or '')
        return jsonify({"ok": True, "message": "数据恢复成功，页面将自动刷新", "csrf_token": session.get("_csrf_token", "")})

    # ── API 列表/统计 ────────────────────────────────────────
    @app.route("/api/cert")
    @login_required
    def api_list_certs() -> Any:
        page = request.args.get("page", 1, type=int)
        per_page = request.args.get("per_page", 20, type=int)
        per_page = min(per_page, 100)
        status_filter = request.args.get("status", "")
        search = request.args.get("search", "").strip()

        certs = load_certs()
        current_username = session.get("username", "")
        users = load_users()
        current_user = next((u for u in users if u["username"] == current_username), None)
        is_admin = current_user.get("role") == "admin" if current_user else False
        if not is_admin:
            certs = [c for c in certs if c.get("created_by") == current_username]

        if status_filter == "expiring":
            certs = [c for c in certs if get_cert_status(c) == "expiring"]
        elif status_filter == "expired":
            certs = [c for c in certs if get_cert_status(c) == "expired"]
        elif status_filter == "normal":
            certs = [c for c in certs if get_cert_status(c) == "normal"]

        if search:
            certs = [c for c in certs if search.lower() in c.get("customer", "").lower()
                     or search.lower() in c.get("domain", "").lower()]

        certs.sort(key=lambda x: calc_days_left(x.get("expire_date", "")))

        total = len(certs)
        start = (page - 1) * per_page
        end = start + per_page
        page_certs = certs[start:end]

        for c in page_certs:
            c["days_left"] = calc_days_left(c.get("expire_date", ""))
            c["status"] = get_cert_status(c, c["days_left"])

        return jsonify({
            "ok": True, "data": page_certs, "total": total,
            "page": page, "per_page": per_page,
            "pages": (total + per_page - 1) // per_page if per_page > 0 else 0,
            "csrf_token": session.get("_csrf_token", "")
        })

    @app.route("/api/stats")
    @login_required
    def api_stats() -> Any:
        certs = load_certs()
        return jsonify(calc_stats(certs))

    # ── 导入预览 ─────────────────────────────────────────────
    @app.route("/api/import-preview", methods=["POST"])
    @login_required
    @admin_required
    def api_import_preview() -> Any:
        """导入预览：接收 JSON/CSV 数据，返回验证结果，不实际导入"""
        if not _check_api_csrf():
            return jsonify({"ok": False, "message": "CSRF验证失败", "csrf_token": session.get("_csrf_token")})
        try:
            # 尝试解析 JSON
            data = request.get_json(silent=True)
            if data is not None:
                if not isinstance(data, list):
                    return jsonify({"ok": False, "message": "JSON 数据必须是数组格式", "csrf_token": session.get("_csrf_token")})
                raw_records = data
            elif "file" in request.files:
                # 上传 CSV 文件
                file = request.files["file"]
                if not file.filename:
                    return jsonify({"ok": False, "message": "未选择文件", "csrf_token": session.get("_csrf_token")})
                raw_records, field_mapping = _parse_csv_file(file)
            else:
                return jsonify({"ok": False, "message": "请提供 JSON 数据或 CSV 文件", "csrf_token": session.get("_csrf_token")})

            # 验证每条记录
            existing_certs = load_certs()
            valid_records = []
            invalid_records = []
            success_count = 0
            fail_count = 0

            for i, item in enumerate(raw_records):
                record = _validate_record(item, i + 1, existing_certs)
                if record["valid"]:
                    valid_records.append(record["data"])
                    success_count += 1
                else:
                    invalid_records.append(record)
                    fail_count += 1

            return jsonify({
                "ok": True,
                "total": len(raw_records),
                "valid_count": success_count,
                "invalid_count": fail_count,
                "valid_records": valid_records,
                "invalid_records": invalid_records,
                "field_mapping": {},
                "csrf_token": session.get("_csrf_token"),
            })
        except Exception as e:
            logger.exception("导入预览失败")
            raise ServiceError(f"预览失败: {str(e)}")

    # ── 实际导入（含预览确认后导入）─────────────────────
    @app.route("/api/import-confirm", methods=["POST"])
    @login_required
    @admin_required
    def api_import_confirm() -> Any:
        """确认导入：从预览数据执行实际导入"""
        if not _check_api_csrf():
            return jsonify({"ok": False, "message": "CSRF验证失败", "csrf_token": session.get("_csrf_token")})
        payload = request.get_json() or {}
        records = payload.get("records", [])
        mode = payload.get("mode", "full")  # full | incremental
        field_mapping = payload.get("field_mapping", {})

        if not records:
            return jsonify({"ok": False, "message": "导入数据为空", "csrf_token": session.get("_csrf_token")})

        existing_certs = load_certs()
        imported = 0
        skipped = 0
        errors: list[str] = []
        new_id = max((c["id"] for c in existing_certs), default=0)

        for i, item in enumerate(records):
            try:
                mapped = _apply_field_mapping(item, field_mapping) if field_mapping else item
                customer = str(mapped.get("customer", "")).strip()
                expire_date = str(mapped.get("expire_date", "")).strip()

                if not customer or not expire_date:
                    errors.append(f"第 {i+1} 条: 缺少必填字段(客户名/到期日期)")
                    continue

                # 日期格式验证
                try:
                    calc_days_left(expire_date)
                except Exception:
                    errors.append(f"第 {i+1} 条: 日期格式无效 '{expire_date}'")
                    continue

                # 重复检查
                is_duplicate = False
                if mode == "incremental":
                    cert_id = mapped.get("cert_id")
                    domain = str(mapped.get("domain", "")).strip()
                    ip = str(mapped.get("ip", "")).strip()
                    if cert_id:
                        is_duplicate = any(c.get("id") == cert_id for c in existing_certs)
                    elif domain and ip:
                        is_duplicate = any(
                            c.get("domain") == domain and c.get("ip") == ip
                            for c in existing_certs
                        )
                    elif domain:
                        is_duplicate = any(c.get("domain") == domain for c in existing_certs)

                if is_duplicate:
                    skipped += 1
                    continue

                new_id += 1
                existing_certs.append({
                    "id": new_id,
                    "customer": customer,
                    "cert_type": str(mapped.get("cert_type", "")).strip(),
                    "domain": str(mapped.get("domain", "")).strip(),
                    "expire_date": expire_date,
                    "note": str(mapped.get("note", "")).strip(),
                    "remind_enabled": bool(mapped.get("remind_enabled", True)),
                    "handled": bool(mapped.get("handled", False)),
                    "responsible_users": mapped.get("responsible_users", []),
                    "created_by": session.get("username", ""),
                    "created_at": datetime.now().strftime("%Y-%m-%d %H:%M"),
                })
                imported += 1
            except Exception as e:
                errors.append(f"第 {i+1} 条: {str(e)}")

        save_certs(existing_certs)
        mode_label = "增量" if mode == "incremental" else "全量"
        write_log(session.get("username", "?"), f"{mode_label}导入确认",
                  f"成功 {imported} 条, 跳过 {skipped} 条, 失败 {len(errors)} 条",
                  "到期项", request.remote_addr or '')
        return jsonify({
            "ok": True,
            "imported": imported,
            "skipped": skipped,
            "errors": errors,
            "csrf_token": session.get("_csrf_token"),
        })

    # ── 数据质量报告 ──────────────────────────────────────
    @app.route("/api/data-quality")
    @login_required
    def api_data_quality() -> Any:
        """数据质量报告：扫描所有证书记录，返回质量问题统计"""
        certs = load_certs()
        now = datetime.now()
        report = {
            "total": len(certs),
            "duplicate_domains": [],
            "missing_fields": {"no_customer": [], "no_domain": [], "no_expire_date": [], "no_remind_enabled": []},
            "expired_records": [],
            "long_unupdated": [],
            "statistics": {},
        }

        # 1. 重复域检测
        domain_map: dict[str, list[int]] = {}
        for c in certs:
            d = str(c.get("domain", "")).strip()
            if d:
                domain_map.setdefault(d, []).append(c["id"])
        for d, ids in domain_map.items():
            if len(ids) > 1:
                report["duplicate_domains"].append({"domain": d, "ids": ids, "count": len(ids)})

        # 2. 缺失字段
        for c in certs:
            cid = c.get("id", "?")
            if not c.get("customer", "").strip():
                report["missing_fields"]["no_customer"].append(cid)
            if not c.get("domain", "").strip():
                report["missing_fields"]["no_domain"].append(cid)
            if not c.get("expire_date", "").strip():
                report["missing_fields"]["no_expire_date"].append(cid)
            if c.get("remind_enabled") is False:
                report["missing_fields"]["no_remind_enabled"].append(cid)

        # 3. 过期未处理
        for c in certs:
            if c.get("handled"):
                continue
            days_left = calc_days_left(c.get("expire_date", ""))
            if days_left < 0:
                report["expired_records"].append({
                    "id": c.get("id"),
                    "customer": c.get("customer", ""),
                    "domain": c.get("domain", ""),
                    "expire_date": c.get("expire_date", ""),
                    "days_expired": abs(days_left),
                })

        # 4. 长期未更新（超过90天未更新）
        for c in certs:
            updated_at = c.get("updated_at", "")
            if not updated_at:
                report["long_unupdated"].append({
                    "id": c.get("id"),
                    "customer": c.get("customer", ""),
                    "reason": "从未更新",
                })
                continue
            try:
                # 尝试多种日期格式
                for fmt in ("%Y-%m-%d %H:%M", "%Y-%m-%dT%H:%M", "%Y-%m-%d"):
                    try:
                        upd_dt = datetime.strptime(updated_at, fmt)
                        delta = (now - upd_dt).days
                        if delta > 90:
                            report["long_unupdated"].append({
                                "id": c.get("id"),
                                "customer": c.get("customer", ""),
                                "updated_at": updated_at,
                                "days_since_update": delta,
                            })
                        break
                    except ValueError:
                        continue
            except Exception:
                pass

        # 统计
        report["statistics"] = {
            "total_records": len(certs),
            "duplicate_domain_groups": len(report["duplicate_domains"]),
            "duplicate_domain_records": sum(d["count"] for d in report["duplicate_domains"]),
            "missing_customer": len(report["missing_fields"]["no_customer"]),
            "missing_domain": len(report["missing_fields"]["no_domain"]),
            "missing_expire_date": len(report["missing_fields"]["no_expire_date"]),
            "disabled_remind": len(report["missing_fields"]["no_remind_enabled"]),
            "expired_unhandled": len(report["expired_records"]),
            "long_unupdated": len(report["long_unupdated"]),
        }

        return jsonify({"ok": True, "report": report})

    # ── CSV 导出（可选列）─────────────────────────────────────
    @app.route("/export/csv")
    @login_required
    def export_csv() -> Response:
        certs = load_certs()
        columns = request.args.getlist("columns", type=str)
        # 默认导出所有标准列
        all_columns = ["customer", "cert_type", "domain", "expire_date", "note",
                       "remind_enabled", "handled", "created_by", "created_at"]
        if columns:
            # 用户指定的列，过滤掉不存在的
            available = [c for c in columns if c in all_columns]
            if not available:
                available = all_columns
        else:
            available = all_columns

        output = io.StringIO()
        writer = csv.writer(output)
        # 中文表头
        header_map = {
            "customer": "客户名称",
            "cert_type": "提醒类型",
            "domain": "域名",
            "expire_date": "到期日期",
            "note": "备注",
            "remind_enabled": "提醒启用",
            "handled": "已处理",
            "created_by": "创建人",
            "created_at": "创建时间",
        }
        writer.writerow([header_map.get(c, c) for c in available])
        for c in certs:
            row = []
            for col in available:
                val = c.get(col, "")
                if isinstance(val, (list, dict)):
                    val = json.dumps(val, ensure_ascii=False)
                elif isinstance(val, bool):
                    val = "是" if val else "否"
                elif val is None:
                    val = ""
                else:
                    val = str(val)
                row.append(val)
            writer.writerow(row)

        buf = io.BytesIO()
        buf.write(output.getvalue().encode("utf-8-sig"))
        buf.seek(0)
        return Response(
            buf.read(),
            mimetype="text/csv",
            headers={"Content-Disposition": "attachment; filename=cert_data_export.csv"},
        )

    # ── CSV 文件上传导入 ──────────────────────────────────────
    @app.route("/api/import-csv", methods=["POST"])
    @login_required
    @admin_required
    def api_import_csv() -> Any:
        """CSV 文件上传导入（含预览）"""
        if not _check_api_csrf():
            return jsonify({"ok": False, "message": "CSRF验证失败", "csrf_token": session.get("_csrf_token")})
        if "file" not in request.files:
            return jsonify({"ok": False, "message": "未上传文件", "csrf_token": session.get("_csrf_token")})
        file = request.files["file"]
        if not file.filename:
            return jsonify({"ok": False, "message": "请选择文件", "csrf_token": session.get("_csrf_token")})

        try:
            raw_records, field_mapping = _parse_csv_file(file)
            # 验证
            existing_certs = load_certs()
            valid_records = []
            invalid_records = []
            success_count = 0
            fail_count = 0

            for i, item in enumerate(raw_records):
                record = _validate_record(item, i + 1, existing_certs)
                if record["valid"]:
                    valid_records.append(record["data"])
                    success_count += 1
                else:
                    invalid_records.append(record)
                    fail_count += 1

            return jsonify({
                "ok": True,
                "total": len(raw_records),
                "valid_count": success_count,
                "invalid_count": fail_count,
                "valid_records": valid_records,
                "invalid_records": invalid_records,
                "field_mapping": field_mapping,
                "csv_headers": list(field_mapping.values()) if field_mapping else [],
                "csrf_token": session.get("_csrf_token"),
            })
        except Exception as e:
            logger.exception("CSV导入预览失败")
            raise ServiceError(f"CSV导入预览失败: {str(e)}")

    # ── 内部辅助函数 ─────────────────────────────────────────
    @staticmethod
    def _detect_encoding(raw_bytes: bytes) -> str:
        """自动检测文本编码（UTF-8 / GBK / GB2312 / latin-1）"""
        # UTF-8 BOM
        if raw_bytes[:3] == b"\xef\xbb\xbf":
            return "utf-8-sig"
        # UTF-8 无 BOM：尝试解码
        try:
            raw_bytes.decode("utf-8")
            return "utf-8"
        except UnicodeDecodeError:
            pass
        # GBK
        try:
            raw_bytes.decode("gbk")
            return "gbk"
        except UnicodeDecodeError:
            pass
        # GB2312
        try:
            raw_bytes.decode("gb2312")
            return "gb2312"
        except UnicodeDecodeError:
            pass
        return "latin-1"

    @staticmethod
    def _detect_delimiter(text: str) -> str:
        """自动检测 CSV 分隔符（逗号 / 制表符 / 分号）"""
        first_line = text.split("\n")[0]
        delimiters = [",", "\t", ";", "|"]
        counts = {d: first_line.count(d) for d in delimiters}
        best = max(counts, key=counts.get)
        if counts[best] == 0:
            return ","
        return best

    @staticmethod
    def _parse_csv_file(file) -> tuple[list[dict[str, str]], dict[str, str]]:
        """解析 CSV 文件，返回 (记录列表, 字段映射)"""
        raw_bytes = file.read()
        encoding = _detect_encoding(raw_bytes)
        text = raw_bytes.decode(encoding)
        delimiter = _detect_delimiter(text)

        reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
        headers = reader.fieldnames or []

        # 智能字段映射：将 CSV 列名映射到标准字段
        field_mapping = {}
        standard_fields = ["customer", "cert_type", "domain", "expire_date", "note",
                           "remind_enabled", "handled", "created_by", "created_at",
                           "id", "responsible_users", "ip"]
        chinese_aliases = {
            "客户名称": "customer", "客户": "customer", "企业名称": "customer",
            "提醒类型": "cert_type", "证书类型": "cert_type", "类型": "cert_type",
            "域名": "domain", "host": "domain", "主机名": "domain",
            "到期日期": "expire_date", "过期日期": "expire_date", "到期时间": "expire_date",
            "expire_date": "expire_date", "expiry_date": "expire_date",
            "备注": "note", "说明": "note", "描述": "note",
            "提醒启用": "remind_enabled", "是否提醒": "remind_enabled", "提醒": "remind_enabled",
            "已处理": "handled", "处理状态": "handled", "状态": "handled",
            "创建人": "created_by", "负责人": "responsible_users",
            "id": "id", "序号": "id",
        }

        for h in headers:
            h_stripped = h.strip()
            if h_stripped in standard_fields:
                field_mapping[h_stripped] = h_stripped
            elif h_stripped in chinese_aliases:
                field_mapping[h_stripped] = chinese_aliases[h_stripped]
            else:
                # 模糊匹配
                hl = h_stripped.lower()
                for alias, std in chinese_aliases.items():
                    if alias.lower() in hl or hl in alias.lower():
                        field_mapping[h_stripped] = std
                        break
                else:
                    field_mapping[h_stripped] = h_stripped  # 保持原名

        records = []
        for row in reader:
            mapped = {}
            for orig_col, std_field in field_mapping.items():
                val = row.get(orig_col, "")
                if val is not None:
                    mapped[std_field] = str(val).strip()
            if mapped:
                records.append(mapped)

        return records, field_mapping

    @staticmethod
    def _validate_record(item: dict, index: int, existing_certs: list[dict]) -> dict:
        """验证单条记录，返回验证结果"""
        customer = str(item.get("customer", "")).strip()
        expire_date = str(item.get("expire_date", "")).strip()
        cert_type = str(item.get("cert_type", "")).strip()
        domain = str(item.get("domain", "")).strip()
        note = str(item.get("note", "")).strip()
        remind_enabled = item.get("remind_enabled")
        handled = item.get("handled")
        responsible_users = item.get("responsible_users", [])
        cert_id = item.get("id") or item.get("cert_id")

        errors = []

        if not customer:
            errors.append(f"第 {index} 条: 缺少客户名称")
        if not expire_date:
            errors.append(f"第 {index} 条: 缺少到期日期")
        else:
            try:
                calc_days_left(expire_date)
            except Exception:
                errors.append(f"第 {index} 条: 日期格式无效 '{expire_date}'")

        # 重复检查
        if customer and expire_date:
            dup = any(
                c.get("customer") == customer and c.get("expire_date") == expire_date
                for c in existing_certs
            )
            if dup:
                errors.append(f"第 {index} 条: 重复记录 (客户名: {customer}, 到期日期: {expire_date})")

        if errors:
            return {"valid": False, "errors": errors, "record_index": index}

        # 解析布尔值
        if remind_enabled is None:
            remind_enabled = True
        elif isinstance(remind_enabled, str):
            remind_enabled = remind_enabled.strip().lower() in ("是", "yes", "true", "1", "y")
        else:
            remind_enabled = bool(remind_enabled)

        if handled is None:
            handled = False
        elif isinstance(handled, str):
            handled = handled.strip().lower() in ("是", "yes", "true", "1", "y")
        else:
            handled = bool(handled)

        if isinstance(responsible_users, str):
            responsible_users = [u.strip() for u in responsible_users.split(",") if u.strip()]

        clean_data = {
            "customer": customer,
            "cert_type": cert_type,
            "domain": domain,
            "expire_date": expire_date,
            "note": note,
            "remind_enabled": remind_enabled,
            "handled": handled,
            "responsible_users": responsible_users,
        }
        if cert_id:
            try:
                clean_data["id"] = int(cert_id)
            except (ValueError, TypeError):
                pass

        return {"valid": True, "data": clean_data, "record_index": index}

    @staticmethod
    def _apply_field_mapping(item: dict, mapping: dict) -> dict:
        """应用字段映射：将原始 CSV 列名映射为标准字段"""
        mapped = {}
        for orig_col, std_field in mapping.items():
            if orig_col in item:
                mapped[std_field] = item[orig_col]
        # 也保留未被映射的原始字段
        for k, v in item.items():
            if k not in mapping:
                mapped[k] = v
        return mapped


def _detect_encoding(raw_bytes: bytes) -> str:
    """模块级编码检测"""
    if raw_bytes[:3] == b"\xef\xbb\xbf":
        return "utf-8-sig"
    try:
        raw_bytes.decode("utf-8")
        return "utf-8"
    except UnicodeDecodeError:
        pass
    try:
        raw_bytes.decode("gbk")
        return "gbk"
    except UnicodeDecodeError:
        pass
    try:
        raw_bytes.decode("gb2312")
        return "gb2312"
    except UnicodeDecodeError:
        pass
    return "latin-1"


def _detect_delimiter(text: str) -> str:
    """模块级分隔符检测"""
    first_line = text.split("\n")[0]
    delimiters = [",", "\t", ";", "|"]
    counts = {d: first_line.count(d) for d in delimiters}
    best = max(counts, key=counts.get)
    if counts[best] == 0:
        return ","
    return best


def _parse_csv_file(file, field_mapping: dict | None = None) -> tuple[list[dict[str, str]], dict[str, str]]:
    """模块级 CSV 解析"""
    raw_bytes = file.read()
    encoding = _detect_encoding(raw_bytes)
    text = raw_bytes.decode(encoding)
    delimiter = _detect_delimiter(text)

    reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
    headers = reader.fieldnames or []

    fm: dict[str, str] = {}
    chinese_aliases = {
        "客户名称": "customer", "客户": "customer", "企业名称": "customer",
        "提醒类型": "cert_type", "证书类型": "cert_type", "类型": "cert_type",
        "域名": "domain", "host": "domain", "主机名": "domain",
        "到期日期": "expire_date", "过期日期": "expire_date", "到期时间": "expire_date",
        "expire_date": "expire_date", "expiry_date": "expire_date",
        "备注": "note", "说明": "note", "描述": "note",
        "提醒启用": "remind_enabled", "是否提醒": "remind_enabled", "提醒": "remind_enabled",
        "已处理": "handled", "处理状态": "handled", "状态": "handled",
        "创建人": "created_by", "负责人": "responsible_users",
        "id": "id", "序号": "id",
    }

    for h in headers:
        h_stripped = h.strip()
        if h_stripped in chinese_aliases:
            fm[h_stripped] = chinese_aliases[h_stripped]
        else:
            hl = h_stripped.lower()
            for alias, std in chinese_aliases.items():
                if alias.lower() in hl or hl in alias.lower():
                    fm[h_stripped] = std
                    break
            else:
                fm[h_stripped] = h_stripped

    records = []
    for row in reader:
        mapped = {}
        for orig_col, std_field in fm.items():
            val = row.get(orig_col, "")
            if val is not None:
                mapped[std_field] = str(val).strip()
        if mapped:
            records.append(mapped)

    return records, fm


def _validate_record(item: dict, index: int, existing_certs: list[dict]) -> dict:
    """模块级记录验证"""
    customer = str(item.get("customer", "")).strip()
    expire_date = str(item.get("expire_date", "")).strip()
    cert_type = str(item.get("cert_type", "")).strip()
    domain = str(item.get("domain", "")).strip()
    note = str(item.get("note", "")).strip()
    remind_enabled = item.get("remind_enabled")
    handled = item.get("handled")
    responsible_users = item.get("responsible_users", [])
    cert_id = item.get("id") or item.get("cert_id")

    errors = []

    if not customer:
        errors.append(f"第 {index} 条: 缺少客户名称")
    if not expire_date:
        errors.append(f"第 {index} 条: 缺少到期日期")
    else:
        try:
            calc_days_left(expire_date)
        except Exception:
            errors.append(f"第 {index} 条: 日期格式无效 '{expire_date}'")

    if customer and expire_date:
        dup = any(
            c.get("customer") == customer and c.get("expire_date") == expire_date
            for c in existing_certs
        )
        if dup:
            errors.append(f"第 {index} 条: 重复记录 (客户名: {customer}, 到期日期: {expire_date})")

    if errors:
        return {"valid": False, "errors": errors, "record_index": index}

    if remind_enabled is None:
        remind_enabled = True
    elif isinstance(remind_enabled, str):
        remind_enabled = remind_enabled.strip().lower() in ("是", "yes", "true", "1", "y")
    else:
        remind_enabled = bool(remind_enabled)

    if handled is None:
        handled = False
    elif isinstance(handled, str):
        handled = handled.strip().lower() in ("是", "yes", "true", "1", "y")
    else:
        handled = bool(handled)

    if isinstance(responsible_users, str):
        responsible_users = [u.strip() for u in responsible_users.split(",") if u.strip()]

    clean_data = {
        "customer": customer,
        "cert_type": cert_type,
        "domain": domain,
        "expire_date": expire_date,
        "note": note,
        "remind_enabled": remind_enabled,
        "handled": handled,
        "responsible_users": responsible_users,
    }
    if cert_id:
        try:
            clean_data["id"] = int(cert_id)
        except (ValueError, TypeError):
            pass

    return {"valid": True, "data": clean_data, "record_index": index}


def _apply_field_mapping(item: dict, mapping: dict) -> dict:
    """模块级字段映射"""
    mapped = {}
    for orig_col, std_field in mapping.items():
        if orig_col in item:
            mapped[std_field] = item[orig_col]
    for k, v in item.items():
        if k not in mapping:
            mapped[k] = v
    return mapped
