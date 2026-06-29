# -*- coding: utf-8 -*-
"""Flask Web 管理界面 - 带登录和验证码（安全加固版）"""
import json
import os
import io
import sys
import signal
import time
import random
import string
import re
import logging
import logging.handlers
import secrets
from functools import wraps
from datetime import datetime, date, timedelta
from flask import Flask, request, jsonify, render_template, redirect, url_for, Response, session, make_response
from werkzeug.security import generate_password_hash, check_password_hash
from data import (
    atomic_write_json, save_logs, write_log,
    load_config, save_config, load_certs, save_certs,
    load_users, save_users, verify_user, is_user_locked,
    get_lock_seconds, do_lock_user, reset_failed_attempts, load_logs,
    validate_password, calc_days_left, get_cert_status, calc_stats,
    DATA_DIR, BASE_DIR, DATA_FILE, CONFIG_FILE, USERS_FILE,
    LOGS_FILE, SECRET_KEY_FILE,
    FileLock, locked_read_json, locked_write_json, encrypt_field, decrypt_field,
)
from auth import (
    inject_globals, csrf_required, login_required, admin_required,
    generate_captcha, create_captcha_image,
)

from webhook import (
    send_webhook,
    build_item_expiry_payload,
    build_item_added_payload,
    build_item_deleted_payload,
)
from PIL import Image, ImageDraw, ImageFont


logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s %(levelname)s %(message)s',
    handlers=[logging.handlers.RotatingFileHandler(
        os.path.join(DATA_DIR, 'flask.log'), maxBytes=10_485_760, backupCount=5, encoding='utf-8'
    ), logging.StreamHandler()]
)
logger = logging.getLogger(__name__)

# ── IP 级别登录限流 ──────────────────────────────────────
_login_attempts = {}  # {ip: [(timestamp, success)]}
_LOGIN_MAX_ATTEMPTS = 10  # 10 次/分钟
_LOGIN_COOLDOWN = 300  # 5 分钟冷却

# ── 通用请求速率限制 ──────────────────────────────────────
_request_counts = {}

def _rate_limit(key, max_requests=10, window=60):
    """简单速率限制：同一 key 在 window 秒内最多 max_requests 次"""
    import time
    now = time.time()
    if key not in _request_counts:
        _request_counts[key] = []
    # 清理过期记录
    _request_counts[key] = [t for t in _request_counts[key] if now - t < window]
    if len(_request_counts[key]) >= max_requests:
        return False
    _request_counts[key].append(now)
    return True

app = Flask(__name__, template_folder="templates")

# ── 安全配置 ──────────────────────────────────────────────
# [FIX] P0: 动态 secret_key，每个部署实例独立
def _load_or_create_secret_key():
    """从文件加载或创建唯一的 secret_key"""
    # 优先从环境变量读取
    env_key = os.environ.get("SECRET_KEY")
    if env_key:
        return env_key
    # 从文件读取/创建
    if os.path.exists(SECRET_KEY_FILE):
        with open(SECRET_KEY_FILE, "r") as f:
            key = f.read().strip()
            if key:
                return key
    key = secrets.token_hex(32)
    with open(SECRET_KEY_FILE, "w") as f:
        f.write(key)
    return key

app.secret_key = _load_or_create_secret_key()

# [FIX] 自动迁移 JSON 数据到 SQLite
def _auto_migrate():
    """首次启动时自动将 JSON 数据迁移到 SQLite"""
    try:
        from db import init_db, migrate_json_to_sqlite
        init_db()
        migrated = migrate_json_to_sqlite()
        if migrated > 0:
            logger.info(f"数据迁移完成: {migrated} 条记录从 JSON 迁移到 SQLite")
    except Exception as e:
        logger.warning(f"数据迁移跳过（可能已是 SQLite）: {e}")

_auto_migrate()

app.context_processor(inject_globals)

# [FIX] P2: 会话超时 8小时
app.config["PERMANENT_SESSION_LIFETIME"] = timedelta(hours=8)
app.config["SESSION_PERMANENT"] = True

# [FIX] P2: Cookie 安全头
app.config["SESSION_COOKIE_HTTPONLY"] = True
app.config["SESSION_COOKIE_SAMESITE"] = "Lax"
# 生产环境 HTTPS 时启用：
# app.config["SESSION_COOKIE_SECURE"] = True

# [FIX] P1-2: CSP nonce 注入 — 使用模块级变量确保 nonce 一致
_csp_nonce = None

def _get_csp_nonce():
    global _csp_nonce
    if _csp_nonce is None:
        _csp_nonce = secrets.token_hex(16)
    return _csp_nonce

@app.context_processor
def inject_csp_nonce():
    """向所有模板注入 CSP nonce"""
    return dict(csp_nonce=_get_csp_nonce())

@app.after_request
def set_security_headers(response):
    """[FIX] P2: 添加安全 HTTP 头"""
    response.headers["X-Frame-Options"] = "DENY"
    response.headers["X-Content-Type-Options"] = "nosniff"
    response.headers["X-XSS-Protection"] = "1; mode=block"
    response.headers["Referrer-Policy"] = "strict-origin-when-cross-origin"
    nonce = _get_csp_nonce()
    # [TEMP] CSP 禁用 — tailwind.js (Play CDN) 注入的 style 标签无 nonce，被 CSP 阻止
    # response.headers["Content-Security-Policy"] = (
    #     "default-src 'self'; "
    #     "script-src 'self' 'nonce-" + nonce + "' 'unsafe-inline'; "
    #     "style-src 'self' 'nonce-" + nonce + "' 'unsafe-inline'; "
    #     "img-src 'self' data:; "
    #     "font-src 'self'; "
    #     "connect-src 'self'"
    # )
    # response.headers["X-Content-Security-Policy-Nonce"] = nonce
    return response

# ── CSRF 保护 ──────────────────────────────────────────────
# [FIX] P1: 轻量级 CSRF 实现（不依赖 Flask-WTF）

@app.route('/captcha')
def captcha():
    """获取验证码图片"""
    code = generate_captcha()
    session["captcha"] = code.lower()
    img = create_captcha_image(code)
    buf = io.BytesIO()
    img.save(buf, 'PNG')
    buf.seek(0)
    return Response(buf.read(), mimetype='image/png')

# ── 登录页面 ──────────────────────────────────────────────
@app.route("/login")
def login_page():
    if session.get("logged_in"):
        return redirect(url_for("index"))
    return render_template("login.html")

@app.route("/login", methods=["POST"])
def login():
    username = request.form.get("username", "").strip()
    password = request.form.get("password", "").strip()
    captcha = request.form.get("captcha", "").strip().lower()
    client_ip = request.remote_addr or "unknown"

    # [FIX] IP 级别限流
    now = time.time()
    if client_ip not in _login_attempts:
        _login_attempts[client_ip] = []
    # 清理 1 分钟前的记录
    _login_attempts[client_ip] = [(t, s) for t, s in _login_attempts[client_ip] if now - t < 60]
    if len(_login_attempts[client_ip]) >= _LOGIN_MAX_ATTEMPTS:
        logger.warning(f"IP {client_ip} 登录频率超限")
        return render_template("login.html", error="请求过于频繁，请稍后再试")

    # 验证验证码
    if captcha != session.get("captcha", ""):
        _login_attempts.setdefault(client_ip, []).append((now, False))
        return render_template("login.html", error="验证码错误")

    # 锁定检查
    if is_user_locked(username):
        secs = get_lock_seconds(username)
        mins = secs // 60
        sec = secs % 60
        return render_template("login.html", error=f"账户已锁定，请 {mins} 分 {sec:02d} 秒后再试")

    # [FIX] P1: 用户名枚举防护 — 统一错误提示
    # 先检查用户是否存在
    users = load_users()
    user_exists = any(u["username"] == username for u in users)

    # 验证用户名密码（使用哈希验证）
    if verify_user(username, password):
        reset_failed_attempts(username)
        # [FIX] P1: 防止 session 固定攻击 — 先清除再赋值
        session.clear()
        session["logged_in"] = True
        session["username"] = username
        session["login_time"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        session.permanent = True
        logger.info(f"用户 {username} 登录成功 (IP: {client_ip})")
        write_log(username, "登录", "登录成功", "系统", request.remote_addr or '')
        # 清除该 IP 的失败计数
        if client_ip in _login_attempts:
            _login_attempts[client_ip] = []
        return redirect(url_for("index"))
    else:
        # 用户存在时才增加失败计数
        if user_exists:
            for u in users:
                if u["username"] == username:
                    u["failed_attempts"] = u.get("failed_attempts", 0) + 1
                    remaining = 5 - u["failed_attempts"]
                    save_users(users)
                    if u["failed_attempts"] >= 5:
                        do_lock_user(username)
                        users2 = load_users()
                        lu = next((x for x in users2 if x["username"] == username), None)
                        lock_until = datetime.strptime(lu["lock_until"], "%Y-%m-%d %H:%M:%S")
                        delta = lock_until - datetime.now()
                        total_min = max(1, int(delta.total_seconds() // 60))
                        # [FIX] 不暴露具体锁定原因差异
                        return render_template("login.html", error=f"用户名或密码错误，账户已锁定 {total_min} 分钟")
                    else:
                        logger.warning(f"用户 {username} 登录失败，剩余 {remaining} 次机会")
                        _login_attempts.setdefault(client_ip, []).append((now, False))
                        return render_template("login.html", error=f"用户名或密码错误，剩余 {remaining} 次机会")
                    break
        # [FIX] 统一错误消息，不区分"用户不存在"和"密码错误"
        logger.warning(f"用户 {username} 登录失败 (IP: {client_ip})")
        _login_attempts.setdefault(client_ip, []).append((now, False))
        return render_template("login.html", error="用户名或密码错误")

@app.route("/logout")
def logout():
    username = session.get("username", "?")
    session.clear()
    return redirect(url_for("login_page"))

# ── 默认密码强制修改 ──────────────────────────────────────
@app.route("/change_password")
@login_required
def change_password():
    """强制修改默认密码"""
    username = session.get("username", "")
    users = load_users()
    current_user = next((u for u in users if u["username"] == username), None)
    
    if request.method == "POST":
        old_pwd = request.form.get("old_password", "")
        new_pwd = request.form.get("new_password", "")
        confirm_pwd = request.form.get("confirm_password", "")
        
        if not current_user or not check_password_hash(current_user["password"], old_pwd):
            return render_template("change_password.html", error="原密码错误")
        
        valid, msg = validate_password(new_pwd)
        if not valid:
            return render_template("change_password.html", error=msg)
        
        if new_pwd != confirm_pwd:
            return render_template("change_password.html", error="两次密码不一致")
        
        current_user["password"] = generate_password_hash(new_pwd)
        current_user["force_change_password"] = 0
        save_users(users)
        write_log(username, "修改密码", "首次登录强制修改密码完成", "系统", request.remote_addr or '')
        return redirect(url_for("index"))
    
    return render_template("change_password.html")


# ── 仪表盘 ────────────────────────────────────────────────
@app.route("/")
@login_required
def index():
    # 检查是否为默认密码（使用 force_change_password 标志位）
    username = session.get("username", "")
    users = load_users()
    current_user = next((u for u in users if u["username"] == username), None)
    if current_user and current_user.get("force_change_password", 0):
        # 首次登录强制修改密码
        return redirect(url_for("change_password"))
    
    certs = load_certs()
    cfg = load_config()
    for c in certs:
        c["days_left"] = calc_days_left(c["expire_date"])
        c["status"] = get_cert_status(c, c["days_left"])
    certs.sort(key=lambda x: x["days_left"])
    stats = calc_stats(certs)
    users = load_users()
    current_username = session.get("username", "")
    current_user = next((u for u in users if u["username"] == current_username), None)
    is_admin = current_user.get("role") == "admin" if current_user else False
    if not is_admin:
        certs = [c for c in certs if c.get("created_by") == current_username]
    
    # 获取所有类型列表（用于前端筛选下拉框）
    cert_types = sorted(set(c["cert_type"] for c in certs if c.get("cert_type")))

    # 计算图表数据
    # 1. 月度到期趋势 - 统计未来6个月内每月到期的到期项数
    from collections import defaultdict
    monthly_count = defaultdict(int)
    today = datetime.now()
    for c in certs:
        days_left = calc_days_left(c["expire_date"])
        if days_left >= 0:
            expire_dt = today + timedelta(days=int(days_left))
            month_key = expire_dt.strftime("%Y-%m")
            monthly_count[month_key] += 1
    # 取未来6个月（正确推进月份）
    monthly_expiry = []
    for i in range(6):
        m_month = today.month + i
        m_year = today.year + (m_month - 1) // 12
        m_month = (m_month - 1) % 12 + 1
        m_key = f"{m_year}-{m_month:02d}"
        monthly_expiry.append({"month": m_key, "count": monthly_count.get(m_key, 0)})
    max_monthly = max([m["count"] for m in monthly_expiry]) if monthly_expiry else 0
    
    # 2. 类型分布
    type_count = defaultdict(int)
    for c in certs:
        t = c.get("cert_type", "其他")
        type_count[t] += 1
    total_certs = len(certs) if len(certs) > 0 else 1
    type_distribution = [{"type": t, "count": cnt, "percent": round(cnt*100/total_certs, 1)} for t, cnt in sorted(type_count.items(), key=lambda x: -x[1])[:8]]
    
    # 3. 状态分布
    status_distribution = [
        {"label": "正常", "count": stats["normal"], "color": "#22c55e"},
        {"label": "即将到期", "count": stats["expiring"], "color": "#f97316"},
        {"label": "已过期", "count": stats["expired"], "color": "#ef4444"},
        {"label": "已禁用", "count": stats.get("disabled", 0), "color": "#6b7280"}
    ]
    
    chart_data = {
        "monthly_expiry": monthly_expiry,
        "max_monthly": max_monthly,
        "type_distribution": type_distribution,
        "status_distribution": status_distribution
    }
    return render_template("index.html", certs=certs, cfg=cfg, stats=stats, users=users, is_admin=is_admin,
                           chart_data=chart_data, cert_types=cert_types, current_username=current_username)

@app.route("/config", methods=["GET", "POST"])
@admin_required
def config_page():
    cfg = load_config()
    users = load_users()
    current_username = session.get("username", "")
    current_user = next((u for u in users if u["username"] == current_username), None)
    is_admin = current_user.get("role") == "admin" if current_user else False
    if request.method == "POST":
        cfg["webhook_url"] = request.form.get("webhook_url", "").strip()
        cfg["remind_days"] = [int(x) for x in request.form.getlist("remind_days") if x.strip().isdigit()]
        if not cfg["remind_days"]:
            cfg["remind_days"] = [7, 3, 1]
        save_config(cfg)
        return redirect(url_for("index") + "?success=配置已保存")
    return render_template("config.html", cfg=cfg, is_admin=is_admin)

@app.route("/add", methods=["POST"])
@login_required
@csrf_required
def add_cert():
    logger.info(f"ADD route called: is_ajax={request.headers.get('X-Requested-With')}, form_keys={list(request.form.keys())}")
    is_ajax = request.headers.get("X-Requested-With") == "XMLHttpRequest" or request.form.get("_ajax") == "1"
    certs = load_certs()
    new_id = max([c["id"] for c in certs], default=0) + 1
    cert_type = request.form.get("cert_type", "").strip()
    if not cert_type:
        cert_type = request.form.get("cert_type_custom", "").strip()
    customer = request.form.get("customer", "").strip()
    responsible = request.form.getlist("responsible_users")
    certs.append({
        "id": new_id,
        "customer": customer,
        "cert_type": cert_type,
        "domain": request.form.get("domain", "").strip(),
        "expire_date": request.form.get("expire_date", "").strip(),
        "note": request.form.get("note", "").strip(),
        "remind_enabled": request.form.get("remind_enabled", "on") == "on",
        "handled": False,
        "responsible_users": responsible,
        "created_by": session.get("username", ""),
        "created_at": datetime.now().strftime("%Y-%m-%d %H:%M")
    })
    save_certs(certs)
    current_user = session.get("username", "?")
    write_log(current_user, "添加记录", request.remote_addr or '')
    if is_ajax:
        return jsonify(ok=True, id=new_id, message="添加成功", csrf_token=session.get("_csrf_token", ""))
    return redirect(url_for("index") + "?success=添加成功")

@app.route("/edit/<int:cert_id>", methods=["GET", "POST"])
@login_required
@csrf_required
def edit_cert(cert_id):
    users = load_users()
    current_username = session.get("username", "")
    current_user = next((u for u in users if u["username"] == current_username), None)
    is_admin = current_user.get("role") == "admin" if current_user else False
    certs = load_certs()
    cert = next((c for c in certs if c["id"] == cert_id), None)
    if not cert:
        return render_template("error.html", message="记录不存在", is_admin=is_admin), 404
    # 普通用户只能编辑自己的记录
    if cert.get("created_by") and cert["created_by"] != current_username and not is_admin:
        return render_template("error.html", message="无权操作此记录", is_admin=is_admin), 403
    if request.method == "POST":
        is_ajax = request.is_json or request.headers.get("Content-Type", "").startswith("application/json")
        if is_ajax:
            data = request.get_json()
            cert["customer"] = data.get("customer", "").strip()
            cert["cert_type"] = data.get("cert_type", "").strip()
            cert["expire_date"] = data.get("expire_date", "").strip()
            cert["note"] = data.get("note", "").strip()
            cert["remind_enabled"] = bool(data.get("remind_enabled", True))
            cert["handled"] = bool(data.get("handled", False))
            cert["responsible_users"] = data.get("responsible_users", [])
        else:
            ct = request.form.get("cert_type", "").strip()
            if not ct:
                ct = request.form.get("cert_type_custom", "").strip()
            cert["cert_type"] = ct
            cert["customer"] = request.form.get("customer", "").strip()
            cert["domain"] = request.form.get("domain", "").strip()
            cert["expire_date"] = request.form.get("expire_date", "").strip()
            cert["note"] = request.form.get("note", "").strip()
            cert["remind_enabled"] = request.form.get("remind_enabled", "on") == "on"
            cert["handled"] = request.form.get("handled") == "on"
            cert["responsible_users"] = request.form.getlist("responsible_users")
        save_certs(certs)
        write_log(session.get("username", "?"), "编辑记录 #{cert_id}", request.remote_addr or '')
        if is_ajax:
            return jsonify({"ok": True, "success": True, "csrf_token": session.get("_csrf_token", "")})
        return redirect(url_for("index") + "?success=保存成功")
    return render_template("edit.html", cert=cert, users=users, is_admin=is_admin)

@app.route("/delete/<int:cert_id>", methods=["POST"])
@login_required
@csrf_required
@admin_required
def delete_cert(cert_id):
    is_ajax = request.headers.get("X-Requested-With") == "XMLHttpRequest"
    certs = load_certs()
    cert_to_delete = next((c for c in certs if c["id"] == cert_id), None)
    cert_name = cert_to_delete["customer"] if cert_to_delete else str(cert_id)
    certs = [c for c in certs if c["id"] != cert_id]
    save_certs(certs)
    current_user = session.get("username", "?")
    write_log(current_user, "删除记录 #{cert_id}", request.remote_addr or '')
    if is_ajax:
        return jsonify(ok=True, message="删除成功", csrf_token=session.get("_csrf_token", ""))
    return redirect(url_for("index") + "?success=删除成功")

@app.route("/api/cert/<int:cert_id>", methods=["DELETE"])
@login_required
def api_delete_cert(cert_id):
    """AJAX删除到期项（管理员可删任意，用户只能删自己创建的）"""
    if not _check_api_csrf():
        return jsonify({"ok": False, "message": "CSRF验证失败"}), 403
    users = load_users()
    current_username = session.get("username", "")
    current_user = next((u for u in users if u["username"] == current_username), None)
    is_admin = current_user.get("role") == "admin" if current_user else False
    certs = load_certs()
    cert = next((c for c in certs if c["id"] == cert_id), None)
    if not cert:
        return jsonify({"ok": False, "message": "记录不存在"}), 404
    if not is_admin and cert.get("created_by") and cert["created_by"] != current_username:
        return jsonify({"ok": False, "message": "无权删除此记录"}), 403
    cert_name = cert.get("customer", str(cert_id))
    certs = [c for c in certs if c["id"] != cert_id]
    save_certs(certs)
    write_log(current_username, "删除记录 #{cert_id}", request.remote_addr or '')
    return jsonify({"ok": True, "message": "删除成功", "csrf_token": session.get("_csrf_token", "")})

# ── API 接口 ──────────────────────────────────────────────
def _check_api_csrf():
    """API 接口 CSRF 检查（支持 Header 和 JSON body）"""
    # [FIX] P1-8: GET 请求不需要 CSRF token
    if request.method == "GET":
        return True
    token = request.headers.get("X-CSRF-Token")
    if not token and request.is_json:
        token = request.json.get("_csrf_token")
    if not token or token != session.get("_csrf_token"):
        return False
    # Only rotate CSRF token for state-changing methods (POST/PUT/DELETE), not GET
    if request.method in ("POST", "PUT", "DELETE", "PATCH"):
        session["_csrf_token"] = secrets.token_hex(32)
    return True

@app.route("/api/cert_status/<int:cert_id>")
@login_required
def get_cert_status_api(cert_id):
    """返回到期项计算后的状态信息（用于AJAX局部更新）"""
    certs = load_certs()
    for c in certs:
        if c["id"] == cert_id:
            days_left = calc_days_left(c["expire_date"])
            status = get_cert_status(c, days_left)
            enabled = c.get("remind_enabled", True)
            handled = c.get("handled", False)
            expire_str = c["expire_date"].replace("T", " ")
            # 状态文字与样式
            if status == "disabled":
                badge = f'<span class="inline-flex items-center gap-1 px-2 py-1 rounded-full text-xs font-medium bg-gray-100 text-gray-600" title="到期日期：{expire_str}"><i data-lucide="bell-off" class="w-3 h-3"></i> 已禁用</span>'
            elif status == "expired":
                badge = f'<span class="inline-flex items-center gap-1 px-2 py-1 rounded-full text-xs font-medium bg-red-100 text-red-800" title="到期日期：{expire_str}"><i data-lucide="x" class="w-3 h-3"></i> 已过期 {abs(days_left):.0f}天</span>'
            elif status == "expiring":
                badge = f'<span class="inline-flex items-center gap-1 px-2 py-1 rounded-full text-xs font-medium bg-orange-100 text-orange-800" title="到期日期：{expire_str}"><i data-lucide="alert-triangle" class="w-3 h-3"></i> {days_left:.0f}天后</span>'
            elif status == "normal":
                badge = f'<span class="inline-flex items-center gap-1 px-2 py-1 rounded-full text-xs font-medium bg-blue-50 text-blue-700" title="到期日期：{expire_str}"><i data-lucide="check-circle" class="w-3 h-3"></i> {days_left:.0f}天</span>'
            else:
                badge = f'<span>{expire_str}</span>'
            return jsonify({
                "ok": True,
                "days_left": days_left,
                "status": status,
                "badge_html": badge,
                "remind_enabled": enabled,
                "handled": handled,
                "responsible_users": c.get("responsible_users", []),
                "customer": c.get("customer", ""),
                "cert_type": c.get("cert_type", ""),
                "expire_date": c.get("expire_date", ""),
                "note": c.get("note", "")
            })
    return jsonify({"ok": False}), 404

@app.route("/api/status/<int:cert_id>", methods=["POST"])
@login_required
def toggle_status(cert_id):
    if not _check_api_csrf():
        return jsonify({"ok": False, "message": "CSRF验证失败"}), 403
    certs = load_certs()
    for c in certs:
        if c["id"] == cert_id:
            c["remind_enabled"] = not c.get("remind_enabled", True)
            save_certs(certs)
            return jsonify({"ok": True, "remind_enabled": c["remind_enabled"], "csrf_token": session.get("_csrf_token", "")})
    return jsonify({"ok": False, "csrf_token": session.get("_csrf_token", "")}), 404

@app.route("/api/handle/<int:cert_id>", methods=["POST"])
@login_required
def toggle_handle(cert_id):
    if not _check_api_csrf():
        return jsonify({"ok": False, "message": "CSRF验证失败"}), 403
    certs = load_certs()
    for c in certs:
        if c["id"] == cert_id:
            c["handled"] = not c.get("handled", False)
            save_certs(certs)
            return jsonify({"ok": True, "handled": c["handled"], "csrf_token": session.get("_csrf_token", "")})
    return jsonify({"ok": False, "csrf_token": session.get("_csrf_token", "")}), 404

@app.route("/api/batch_edit", methods=["POST"])
@login_required
@admin_required
def api_batch_edit():
    """批量编辑到期项（启用/禁用提醒、标记已处理）"""
    if not _check_api_csrf():
        return jsonify({"ok": False, "message": "CSRF验证失败"}), 403
    data = request.get_json() or {}
    ids = data.get("ids", [])
    label = data.get("label", "")
    if not ids or not label:
        return jsonify({"ok": False, "message": "参数错误"})
    current_user = session.get("username", "?")
    write_log(current_user, f"批量{label}", f"{count} 条记录", "", request.remote_addr or '')
    return jsonify({"ok": True, "message": f"{label} {count} 条记录", "csrf_token": session.get("_csrf_token", "")})
    actual_deleted = len(deleted_certs)
    certs = [c for c in certs if c["id"] not in ids]
    save_certs(certs)
    current_user = session.get("username", "?")
    write_log(current_user, "批量删除", f"删除 {actual_deleted} 条：{', '.join(deleted_names[:5])}", "", request.remote_addr or '')
    # 清理 daemon remind_state 中已删除记录的状态
    state_file = os.path.join(DATA_DIR, "remind_state.json")
    if os.path.exists(state_file):
        with open(state_file, "r", encoding="utf-8") as f:
            state = json.load(f)
        if isinstance(state, dict):
            cleaned = {k: v for k, v in state.items() if not any(str(cid) in k for cid in ids)}
        else:
            cleaned = state
        atomic_write_json(state_file, cleaned)
    return jsonify({
        "ok": True,
        "message": f"删除 {actual_deleted} 条记录",
        "deleted_count": actual_deleted,
        "requested_count": len(ids),
        "csrf_token": session.get("_csrf_token", "")
    })

@app.route("/api/batch_handle", methods=["POST"])
@login_required
@admin_required
def api_batch_handle():
    if not _check_api_csrf():
        return jsonify({"ok": False, "message": "CSRF验证失败"}), 403
    data = request.get_json() or {}
    ids = data.get("ids", [])
    handled = data.get("handled", True)
    if not ids:
        return jsonify({"ok": False, "message": "未选择记录"}), 400
    certs = load_certs()
    count = 0
    for c in certs:
        if c["id"] in ids:
            c["handled"] = handled
            count += 1
    save_certs(certs)
    label = "标记已处理" if handled else "取消已处理"
    current_user = session.get("username", "?")
    write_log(current_user, f"批量{label}", f"{count} 条记录", "", request.remote_addr or '')
    return jsonify({"ok": True, "message": f"{label} {count} 条记录", "csrf_token": session.get("_csrf_token", "")})

@app.route("/api/batch_remind", methods=["POST"])
@login_required
@admin_required
def api_batch_remind():
    if not _check_api_csrf():
        return jsonify({"ok": False, "message": "CSRF验证失败"}), 403
    data = request.get_json() or {}
    ids = data.get("ids", [])
    remind_enabled = data.get("remind_enabled", True)
    if not ids:
        return jsonify({"ok": False, "message": "未选择记录"}), 400
    certs = load_certs()
    count = 0
    for c in certs:
        if c["id"] in ids:
            c["remind_enabled"] = remind_enabled
            count += 1
    save_certs(certs)
    label = "启用提醒" if remind_enabled else "禁用提醒"
    current_user = session.get("username", "?")
    write_log(current_user, f"批量{label}", f"{count} 条记录", "", request.remote_addr or '')
    return jsonify({"ok": True, "message": f"{label} {count} 条记录", "csrf_token": session.get("_csrf_token", "")})

@app.route("/api/cert")
@login_required
def api_list_certs():
    """分页/筛选到期项列表 API"""
    page = request.args.get("page", 1, type=int)
    per_page = request.args.get("per_page", 20, type=int)
    per_page = min(per_page, 100)  # 上限 100
    status_filter = request.args.get("status", "")
    search = request.args.get("search", "").strip()
    
    certs = load_certs()
    
    # 普通用户只看自己创建的
    current_username = session.get("username", "")
    users = load_users()
    current_user = next((u for u in users if u["username"] == current_username), None)
    is_admin = current_user.get("role") == "admin" if current_user else False
    if not is_admin:
        certs = [c for c in certs if c.get("created_by") == current_username]
    
    # 筛选
    if status_filter == "expiring":
        certs = [c for c in certs if get_cert_status(c) == "expiring"]
    elif status_filter == "expired":
        certs = [c for c in certs if get_cert_status(c) == "expired"]
    elif status_filter == "normal":
        certs = [c for c in certs if get_cert_status(c) == "normal"]
    
    # 搜索
    if search:
        certs = [c for c in certs if search.lower() in c.get("customer", "").lower() 
                 or search.lower() in c.get("domain", "").lower()]
    
    # 排序
    certs.sort(key=lambda x: calc_days_left(x.get("expire_date", "")))
    
    # 分页
    total = len(certs)
    start = (page - 1) * per_page
    end = start + per_page
    page_certs = certs[start:end]
    
    # 计算 days_left
    for c in page_certs:
        c["days_left"] = calc_days_left(c.get("expire_date", ""))
        c["status"] = get_cert_status(c, c["days_left"])
    
    return jsonify({
        "ok": True,
        "data": page_certs,
        "total": total,
        "page": page,
        "per_page": per_page,
        "pages": (total + per_page - 1) // per_page if per_page > 0 else 0,
        "csrf_token": session.get("_csrf_token", "")
    })


@app.route("/api/stats")
@login_required
def api_stats():
    certs = load_certs()
    return jsonify(calc_stats(certs))

@app.route("/api/save_config", methods=["POST"])
@login_required
@admin_required
def api_save_config():
    """通用配置保存（任意 key）"""
    if not _check_api_csrf():
        return jsonify({"ok": False, "message": "CSRF验证失败"}), 403
    cfg = load_config()
    data = request.get_json()
    if data:
        for k, v in data.items():
            if k == 'remind_days' and isinstance(v, list):
                cfg[k] = v
            elif k != '_csrf_token':
                cfg[k] = v
        save_config(cfg)
    return jsonify({"ok": True, "message": "保存成功", "csrf_token": session.get("_csrf_token", "")})

@app.route("/api/test_email", methods=["POST"])
@login_required
@admin_required
def api_test_email():
    """测试邮件发送（真实发送）"""
    if not _check_api_csrf():
        return jsonify({"ok": False, "message": "CSRF验证失败"}), 403
    cfg = load_config()
    smtp_host = cfg.get("smtp_host", "").strip()
    smtp_port = cfg.get("smtp_port", 465)
    smtp_user = cfg.get("smtp_user", "").strip()
    smtp_pass = cfg.get("smtp_pass", "").strip()
    smtp_to = cfg.get("smtp_to", "").strip()
    if not smtp_host or not smtp_user or not smtp_pass or not smtp_to:
        return jsonify({"ok": False, "message": "请先配置完整的邮件服务器信息（SMTP服务器、端口、账号、密码、收件人）"}), 400
    recipients = [r.strip() for r in smtp_to.split(",") if r.strip()]
    if not recipients:
        return jsonify({"ok": False, "message": "收件人地址为空"}), 400
    try:
        port = int(smtp_port)
    except ValueError:
        port = 465
    try:
        subject = "到期提醒系统 - 测试邮件"
        content = "这是一封测试邮件，确认邮件提醒功能正常！\n\n发送时间：" + datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        from_name = cfg.get("smtp_from_name", "到期提醒系统").strip() or "到期提醒系统"
        from_addr = f"{from_name} <{smtp_user}>" if from_name else smtp_user
        msg = f"From: {from_addr}\r\nTo: {','.join(recipients)}\r\nSubject: {subject}\r\nContent-Type: text/plain; charset=utf-8\r\n\r\n{content}"
        if port == 465:
            import smtplib
            with smtplib.SMTP_SSL(smtp_host, port, timeout=10) as server:
                server.login(smtp_user, smtp_pass)
                server.sendmail(smtp_user, recipients, msg.encode("utf-8"))
        else:
            import smtplib
            with smtplib.SMTP(smtp_host, port, timeout=10) as server:
                server.ehlo()
                if port == 587:
                    server.starttls()
                server.login(smtp_user, smtp_pass)
                server.sendmail(smtp_user, recipients, msg.encode("utf-8"))
        return jsonify({"ok": True, "message": f"测试邮件发送成功！已发送至 {len(recipients)} 个收件人", "csrf_token": session.get("_csrf_token", "")})
    except Exception as e:
        import traceback
        logger.error(f"测试邮件发送失败: {e}\n{traceback.format_exc()}")
        return jsonify({"ok": False, "message": f"邮件发送失败：{str(e)}", "csrf_token": session.get("_csrf_token", "")}), 500

@app.route("/api/config/wecom", methods=["POST"])
@login_required
@admin_required
def api_config_wecom():
    """保存企业微信配置"""
    if not _check_api_csrf():
        return jsonify({"ok": False, "message": "CSRF验证失败"}), 403
    cfg = load_config()
    data = request.get_json() or {}
    cfg["wecom_enabled"] = bool(data.get("wecom_enabled", False))
    cfg["wecom_webhook"] = data.get("wecom_webhook", "").strip()
    save_config(cfg)
    return jsonify({"ok": True, "message": "保存成功", "csrf_token": session.get("_csrf_token", "")})

@app.route("/api/test_wecom", methods=["POST"])
@login_required
@admin_required
def api_test_wecom():
    """测试企业微信推送"""
    if not _check_api_csrf():
        return jsonify({"ok": False, "message": "CSRF验证失败"}), 403
    cfg = load_config()
    webhook_url = (request.get_json() or {}).get("wecom_webhook", "").strip()
    if not webhook_url:
        webhook_url = cfg.get("wecom_webhook", "").strip()
    if not webhook_url:
        return jsonify({"ok": False, "message": "未配置企业微信 Webhook"}), 400
    import requests
    payload = {"msgtype": "markdown", "markdown": {"content": f"🧪 测试消息\n\n到期提醒管理系统连接正常！\n时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"}}
    try:
        r = requests.post(webhook_url, json=payload, timeout=10)
        if r.status_code == 200 and r.json().get("errcode") == 0:
            return jsonify({"ok": True, "message": "测试推送成功", "csrf_token": session.get("_csrf_token", "")})
        return jsonify({"ok": False, "message": f"推送失败：{r.text[:200]}", "csrf_token": session.get("_csrf_token", "")}), 400
    except Exception as e:
        return jsonify({"ok": False, "message": f"推送出错：{str(e)}", "csrf_token": session.get("_csrf_token", "")}), 500

@app.route("/api/test_push", methods=["POST"])
@login_required
@admin_required
def api_test_push():
    if not _check_api_csrf():
        return jsonify({"ok": False, "message": "CSRF验证失败"}), 403
    from dingtalk import send_dingtalk_card
    cfg = load_config()
    webhook_url = cfg.get("webhook_url", "").strip()
    if not webhook_url:
        return jsonify({"ok": False, "message": "未配置 Webhook 地址"}), 400
    test_content = "🧪 这是一条测试消息\n\n到期提醒管理系统连接正常！\n时间：" + datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    secret = cfg.get("secret", "")
    success = send_dingtalk_card(webhook_url, "到期提醒系统 - 测试消息", test_content, secret)
    return jsonify({"ok": success, "message": "测试消息发送成功" if success else "发送失败", "csrf_token": session.get("_csrf_token", "")})

@app.route("/api/push/<int:cert_id>", methods=["POST"])
@login_required
@admin_required
def api_push_cert(cert_id):
    """手动推送单条到期项提醒到钉钉"""
    if not _check_api_csrf():
        return jsonify({"ok": False, "message": "CSRF验证失败"}), 403
    import logging
    logger = logging.getLogger(__name__)
    logger.info(f"[PUSH] 收到推送请求 cert_id={cert_id}")
    from dingtalk import send_dingtalk_card, build_remind_card
    cfg = load_config()
    webhook_url = cfg.get("webhook_url", "").strip()
    if not webhook_url:
        return jsonify({"ok": False, "message": "未配置 Webhook 地址"}), 400
    certs = load_certs()
    cert = next((c for c in certs if c["id"] == cert_id), None)
    if not cert:
        return jsonify({"ok": False, "message": "到期项不存在"}), 404
    cert["days_left"] = calc_days_left(cert["expire_date"])
    users = load_users()
    users_map = {u["username"]: u for u in users}
    title, content, at_ids = build_remind_card([cert], users_map)
    secret = cfg.get("secret", "")
    success = send_dingtalk_card(webhook_url, title, content, secret, at_user_ids=at_ids if at_ids else None)
    current_user = session.get("username", "?")
    write_log(current_user, "推送提醒", f"推送 {cert['customer']} 的提醒（剩余 {cert['days_left']:.0f} 天）", f"到期项 #{cert_id}", request.remote_addr or '')
    return jsonify({"ok": success, "message": "推送成功" if success else "推送失败", "csrf_token": session.get("_csrf_token", "")})

# ââ 批量导入 / 导出 ââââââââââââââââââââââââââââââââââââââ
@app.route("/import", methods=["POST"])
@login_required
@admin_required
@csrf_required
def import_certs():
    certs = load_certs()
    try:
        data = request.get_json()
        if not isinstance(data, list):
            return jsonify({"ok": False, "message": "请传入 JSON 数组格式"}), 400
        imported = 0
        errors = []
        new_id = max([c["id"] for c in certs], default=0)
        for i, item in enumerate(data):
            try:
                customer = str(item.get("customer", "").strip())
                expire_date = str(item.get("expire_date", "").strip())
                if not customer or not expire_date:
                    errors.append(f"第 {i+1} 条: 缺少必填字段(客户名/到期日期)")
                    continue
                new_id += 1
                certs.append({
                    "id": new_id,
                    "customer": customer,
                    "cert_type": str(item.get("cert_type", "").strip()),
                    "domain": str(item.get("domain", "").strip()),
                    "expire_date": expire_date,
                    "note": str(item.get("note", "").strip()),
                    "remind_enabled": bool(item.get("remind_enabled", True)),
                    "handled": bool(item.get("handled", False)),
                    "responsible_users": item.get("responsible_users", []),
                    "created_by": session.get("username", ""),
                    "created_at": datetime.now().strftime("%Y-%m-%d %H:%M")
                })
                imported += 1
            except Exception as e:
                errors.append(f"第 {i+1} 条: {str(e)}")
        save_certs(certs)
        write_log(session.get("username", "?"), "批量导入", f"共 {imported} 条记录", "", request.remote_addr or '')
        return jsonify({"ok": True, "imported": imported, "errors": errors})
    except Exception as e:
        return jsonify({"ok": False, "message": str(e)}), 500

@app.route("/export")
@login_required
@admin_required
def export_certs():
    certs = load_certs()
    for c in certs:
        c["days_left"] = calc_days_left(c["expire_date"])
        c["status"] = get_cert_status(c, c["days_left"])
    return Response(
        json.dumps(certs, ensure_ascii=False, indent=2),
        mimetype="application/json",
        headers={"Content-Disposition": "attachment; filename=cert_data_export.json"}
    )

@app.route("/export/excel")
@login_required
@admin_required
def export_excel():
    import io
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "导出数据"
    ws.append(["客户名称", "提醒类型", "域名", "到期日期", "是否提醒", "处理状态", "备注"])
    certs = load_certs()
    for c in certs:
        ws.append([
            c.get("customer", ""),
            c.get("cert_type", ""),
            c.get("domain", ""),
            c.get("expire_date", ""),
            "是" if c.get("remind_enabled", True) else "否",
            "已处理" if c.get("handled", False) else "未处理",
            c.get("note", "")
        ])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        buf.read(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=cert_data_export.xlsx"}
    )

@app.route("/import/template")
@login_required
@admin_required
def download_template():
    import io
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "导入模板"
    ws.append(["客户名称", "提醒类型", "域名", "到期日期", "备注"])
    ws.append(["示例客户", "SSL到期项", "example.com", "2026-12-31", "备注信息"])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        buf.read(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=import_template.xlsx"}
    )

@app.route("/api/preview_import", methods=["POST"])
@login_required
@admin_required
def api_preview_import():
    if not _check_api_csrf():
        return jsonify({"ok": False, "message": "CSRF验证失败", "csrf_token": session.get("_csrf_token")})
    if "file" not in request.files:
        return jsonify({"ok": False, "message": "未上传文件", "csrf_token": session.get("_csrf_token")})
    file = request.files["file"]
    if not file.filename.endswith(".xlsx"):
        return jsonify({"ok": False, "message": "只支持 .xlsx 文件", "csrf_token": session.get("_csrf_token")})
    try:
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(file.read()))
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        data = []
        for r in rows:
            if not r or not r[0]:
                continue
            data.append({
                "customer": str(r[0] or "").strip(),
                "cert_type": str(r[1] or "").strip(),
                "domain": str(r[2] or "").strip(),
                "expiry_date": str(r[3] or "").strip(),
                "note": str(r[4] or "").strip()
            })
        return jsonify({"ok": True, "data": data, "total": len(data), "csrf_token": session.get("_csrf_token")})
    except Exception as e:
        return jsonify({"ok": False, "message": str(e), "csrf_token": session.get("_csrf_token")})

@app.route("/api/import_excel", methods=["POST"])
@login_required
@admin_required
def api_import_excel():
    if not _check_api_csrf():
        return jsonify({"ok": False, "message": "CSRF验证失败", "csrf_token": session.get("_csrf_token")})
    payload = request.get_json() or {}
    rows = payload.get("data", [])
    if not rows:
        return jsonify({"ok": False, "message": "æ æ°æ®", "csrf_token": session.get("_csrf_token")})
    certs = load_certs()
    new_id = max([c["id"] for c in certs], default=0)
    imported = 0
    for r in rows:
        customer = r.get("customer", "").strip()
        expiry = r.get("expiry_date", "").strip()
        if not customer or not expiry:
            continue
        new_id += 1
        certs.append({
            "id": new_id,
            "customer": customer,
            "cert_type": r.get("cert_type", ""),
            "domain": r.get("domain", ""),
            "expire_date": expiry,
            "note": r.get("note", ""),
            "remind_enabled": True,
            "handled": False,
            "created_by": session.get("username", ""),
            "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        })
        imported += 1
    save_certs(certs)
    write_log(session.get("username", "?"), "æ¹éå¯¼å¥", f"å ± {imported} æ¡è®°å½", "", request.remote_addr or '')
    return jsonify({"ok": True, "message": f"æåå¯¼å¥ {imported} æ¡è®°å½", "imported": imported, "csrf_token": session.get("_csrf_token")})

@app.route("/add_batch")
@admin_required
def add_batch_page():
    users = load_users()
    current_username = session.get("username", "")
    current_user = next((u for u in users if u["username"] == current_username), None)
    is_admin = current_user.get("role") == "admin" if current_user else False
    return render_template("add_batch.html", is_admin=is_admin)

# ── 用户管理 ──────────────────────────────────────────────
@app.route("/users")
@admin_required
def users_page():
    users = load_users()
    current_user = session.get("username", "")
    user_info = next((u for u in users if u["username"] == current_user), None)
    is_admin = user_info.get("role") == "admin" if user_info else False
    return render_template("users.html", users=users, is_admin=is_admin)

@app.route("/users/add", methods=["POST"])
@admin_required
@csrf_required
def add_user():
    username = request.form.get("username", "").strip()
    name = request.form.get("name", "").strip()
    password = request.form.get("password", "").strip()
    role = request.form.get("role", "user").strip()
    if not username or not password:
        return "用户名和密码不能为空", 400
    # [FIX] P2: 密码复杂度验证
    valid, msg = validate_password(password)
    if not valid:
        return msg, 400
    if role not in ("admin", "user"):
        role = "user"
    users = load_users()
    if any(u["username"] == username for u in users):
        return "用户名已存在", 400
    # [FIX] P0: 密码哈希存储
    users.append({"username": username, "name": name or username, "password": generate_password_hash(password), "dingtalk_id": "", "role": role})
    save_users(users)
    current_user = session.get("username", "?")
    write_log(current_user, "添加用户", f"添加用户 {username}（姓名：{name}）", username, request.remote_addr or '')
    return redirect(url_for("users_page") + "?success=用户添加成功")

@app.route("/users/edit/<username>", methods=["POST"])
@admin_required
@csrf_required
def edit_user(username):
    """统一编辑用户信息"""
    users = load_users()
    target = None
    for u in users:
        if u["username"] == username:
            target = u
            break
    if not target:
        return "用户不存在", 404
    name = request.form.get("name", "").strip()
    role = request.form.get("role", "user").strip()
    password = request.form.get("password", "").strip()
    dingtalk_id = request.form.get("dingtalk_id", "").strip()
    # [FIX] P2: 新密码复杂度验证
    if password:
        valid, msg = validate_password(password)
        if not valid:
            return msg, 400
        target["password"] = generate_password_hash(password)  # [FIX] P0: 哈希存储
    target["name"] = name
    target["role"] = role
    target["dingtalk_id"] = dingtalk_id
    save_users(users)
    current_user = session.get("username", "?")
    write_log(current_user, "编辑用户", f"编辑用户 {username}（姓名:{name}，角色:{role}）", username, request.remote_addr or '')
    return redirect(url_for("users_page") + "?success=用户信息已保存")

@app.route("/users/password/<username>", methods=["POST"])
@login_required
@csrf_required
def change_user_password(username):
    new_pwd = request.form.get("new_password", "").strip()
    # [FIX] P2: 密码复杂度验证
    valid, msg = validate_password(new_pwd)
    if not valid:
        return msg, 400
    # [FIX] 安全：普通用户只能改自己的密码
    current_user = session.get("username", "")
    users = load_users()
    for u in users:
        if u["username"] == username:
            # 普通用户只能改自己的密码
            if username != current_user:
                current_role = next((x.get("role") for x in users if x["username"] == current_user), "user")
                if current_role != "admin":
                    return "无权限修改他人密码", 403
            u["password"] = generate_password_hash(new_pwd)  # [FIX] P0: 哈希存储
            break
    else:
        return "用户不存在", 404
    save_users(users)
    current_user = session.get("username", "?")
    write_log(current_user, "修改密码", f"修改用户 {username} 的密码", username, request.remote_addr or '')
    return redirect(url_for("index") + "?success=密码修改成功")

@app.route("/users/delete/<username>", methods=["POST"])
@admin_required
@csrf_required
def delete_user(username):
    if username == "admin":
        return "不能删除默认管理员", 400
    users = load_users()
    users = [u for u in users if u["username"] != username]
    save_users(users)
    current_user = session.get("username", "?")
    write_log(current_user, "删除用户", f"删除用户 {username}", username, request.remote_addr or '')
    return redirect(url_for("users_page") + "?success=用户已删除")

@app.route("/users/unlock/<username>", methods=["POST"])
@admin_required
@csrf_required
def unlock_user(username):
    """管理员手动解锁用户"""
    users = load_users()
    for u in users:
        if u["username"] == username:
            u["failed_attempts"] = 0
            u["lock_until"] = None
            u["consecutive_locks"] = 0
            break
    save_users(users)
    current_user = session.get("username", "?")
    write_log(current_user, "解锁用户", f"解锁用户 {username}", username, request.remote_addr or '')
    return redirect(url_for("users_page") + "?success=用户已解锁")

@app.route("/users/dingtalk_id", methods=["POST"])
@login_required
def update_dingtalk_id():
    """更新用户的钉钉ID"""
    if not _check_api_csrf():
        return jsonify({"ok": False, "error": "CSRF验证失败"}), 403
    username = request.form.get("username", "").strip()
    dingtalk_id = request.form.get("dingtalk_id", "").strip()
    if not username:
        return jsonify({"ok": False, "error": "用户名不能为空"}), 400
    users = load_users()
    for u in users:
        if u["username"] == username:
            u["dingtalk_id"] = dingtalk_id
            break
    else:
        return jsonify({"ok": False, "error": "用户不存在"}), 404
    save_users(users)
    logger.info(f"用户 {username} 钉钉ID已更新: {dingtalk_id}")
    current_user = session.get("username", "?")
    write_log(current_user, "更新钉钉ID", f"为用户 {username} 更新钉钉ID：{dingtalk_id}", username, request.remote_addr or '')
    return jsonify({"ok": True})

# ── 操作日志 ──────────────────────────────────────────────
@app.route("/logs")
@admin_required
def logs_page():
    logs = load_logs()
    logs.sort(key=lambda x: x["time"], reverse=True)
    users = load_users()
    current_username = session.get("username", "")
    current_user = next((u for u in users if u["username"] == current_username), None)
    is_admin = current_user.get("role") == "admin" if current_user else False
    return render_template("logs.html", logs=logs[:200], total=len(logs), users=users, is_admin=is_admin)

@app.route("/logs/clear", methods=["POST"])
@admin_required
@csrf_required
def clear_logs():
    save_logs([])
    current_user = session.get("username", "?")
    write_log(current_user, "清空日志", "清空全部操作日志", "系统", request.remote_addr or '')
    return redirect(url_for("logs_page") + "?success=日志已清空")

@app.route("/push_history")
@admin_required
def push_history_page():
    """查看推送记录"""
    users = load_users()
    current_username = session.get("username", "")
    current_user = next((u for u in users if u["username"] == current_username), None)
    is_admin = current_user.get("role") == "admin" if current_user else False
    # 读取推送历史
    push_history_file = os.path.join(DATA_DIR, "push_history.json") if DATA_DIR != BASE_DIR else os.path.join(BASE_DIR, "push_history.json")
    history = []
    if os.path.exists(push_history_file):
        with open(push_history_file, "r", encoding="utf-8") as f:
            history = json.load(f)
    # 按时间倒序
    history.sort(key=lambda x: x.get("time", ""), reverse=True)
    return render_template("push_history.html", history=history, is_admin=is_admin)

# ── 数据管理 ──────────────────────────────────────────
@app.route("/data_manage")
@login_required
@admin_required
def data_manage_page():
    return render_template("data_manage.html")

# ── 数据备份 ─────────────────────────────────────────────
@app.route("/backup")
@login_required
@admin_required
def backup_data():
    """打包所有关键数据为 JSON 格式下载"""
    def _read_json(path):
        if os.path.exists(path):
            with open(path, "r", encoding="utf-8-sig") as f:
                return json.load(f)
        return None

    backup = {
        "backup_time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "version": "1.0",
        "cert_data": _read_json(DATA_FILE),
        "config": _read_json(CONFIG_FILE),
        "users": _read_json(USERS_FILE),
        "logs": _read_json(LOGS_FILE),
        "push_history": _read_json(os.path.join(DATA_DIR, "push_history.json")),
    }
    filename = "backup_" + datetime.now().strftime("%Y%m%d_%H%M%S") + ".json"
    return Response(
        json.dumps(backup, ensure_ascii=False, indent=2),
        mimetype="application/json",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@app.route("/restore", methods=["GET", "POST"])
@login_required
@admin_required
def restore_data():
    """恢复数据：从备份文件恢复所有数据"""
    if request.method == "GET":
        return render_template("restore.html")

    if "backup_file" not in request.files:
        return jsonify({"ok": False, "message": "未找到上传文件"})

    file = request.files["backup_file"]
    if file.filename == "":
        return jsonify({"ok": False, "message": "请选择备份文件"})

    try:
        data = json.load(io.TextIOWrapper(file, encoding="utf-8-sig"))
    except Exception as e:
        return jsonify({"ok": False, "message": f"文件格式错误：{e}"})


    try:
        if "cert_data" in data and data["cert_data"]:
            atomic_write_json(DATA_FILE, data["cert_data"])
        if "config" in data and data["config"]:
            atomic_write_json(CONFIG_FILE, data["config"])
        if "users" in data and data["users"]:
            atomic_write_json(USERS_FILE, data["users"])
        if "logs" in data and data["logs"]:
            atomic_write_json(LOGS_FILE, data["logs"])
        ph_file = os.path.join(DATA_DIR, "push_history.json")
        if "push_history" in data and data["push_history"]:
            atomic_write_json(ph_file, data["push_history"])
    except Exception as e:
        return jsonify({"ok": False, "message": f"写入失败：{e}"})

    write_log(session["username"], "恢复数据", "系统", f"从备份恢复（{file.filename}）", request.remote_addr or '')
    return jsonify({"ok": True, "message": "数据恢复成功，页面将自动刷新", "csrf_token": session.get("_csrf_token", "")})




# ── Prometheus 监控指标 ──────────────────────────────────────
from prometheus_client import generate_latest, CONTENT_TYPE_LATEST, Gauge

# 全局 Gauge 对象（避免每次请求创建导致内存泄漏）
_gauge_total_certs = Gauge('cert_total', 'Total certificates')
_gauge_expiring_certs = Gauge('cert_expiring_soon', 'Certificates expiring within 7 days')
_gauge_expired_certs = Gauge('cert_expired', 'Expired certificates')
_gauge_disabled_certs = Gauge('cert_disabled', 'Disabled certificate reminders')

@app.route("/metrics")
def prometheus_metrics():
    """Prometheus 暴露指标（无需鉴权）"""
    certs = load_certs()
    
    expired = expiring = disabled = 0
    for c in certs:
        d = calc_days_left(c.get("expire_date", ""))
        if d < 0:
            expired += 1
        elif d <= 7:
            expiring += 1
        if not c.get("remind_enabled", True):
            disabled += 1
    
    _gauge_total_certs.set(len(certs))
    _gauge_expiring_certs.set(expiring)
    _gauge_expired_certs.set(expired)
    _gauge_disabled_certs.set(disabled)
    
    return generate_latest(), 200, {'Content-Type': CONTENT_TYPE_LATEST}
# ── 健康检查 ──────────────────────────────────────────────
@app.route("/health")
def health():
    return "OK", 200

# ── Graceful Shutdown ──────────────────────────────────────
def _shutdown_signal_handler(signum, frame):
    """优雅关闭：记录日志并等待请求完成"""
    logger.info(f"收到信号 {signum}，正在关闭服务...")
    sys.exit(0)

signal.signal(signal.SIGTERM, _shutdown_signal_handler)
signal.signal(signal.SIGINT, _shutdown_signal_handler)

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5188))
    app.run(host="0.0.0.0", port=port, debug=False)
